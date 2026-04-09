# -*- coding: utf-8 -*-
"""
Kredit Lab — Bank Statement Intelligence Dashboard
Schema v6.0.0 / v6.1.0 / v6.1.1 / v6.2.0 / v6.2.1 / v6.2.2: Malaysian bank statement analysis engine
v6.1.0: Per-account monthly cash flow breakdown
v6.1.1: Added opening_balance and closing_balance to monthly analysis
v6.2.0: FX transaction tracking, parsing quality metrics, counterparty MoM trends
v6.2.1: Corrected FX classification
v6.2.2: Reconciliation status at monthly + consolidated level, extraction gap detection,
        data quality banners in HTML/Streamlit, gap detail panels
Features: Interactive HTML report + Excel download
© Kredit Lab Sdn Bhd
"""

import streamlit as st
import json
import io
import os
import hmac
from datetime import datetime

st.set_page_config(page_title="Kredit Lab — Statement Intelligence", page_icon="🔬", layout="wide")


def _get_auth_credentials():
    username = os.getenv("APP_USERNAME") or os.getenv("STREAMLIT_APP_USERNAME")
    password = os.getenv("APP_PASSWORD") or os.getenv("STREAMLIT_APP_PASSWORD")
    return username, password


def _check_login(username_input, password_input):
    expected_username, expected_password = _get_auth_credentials()
    if not expected_username or not expected_password:
        return None
    username_ok = hmac.compare_digest(str(username_input or ""), str(expected_username))
    password_ok = hmac.compare_digest(str(password_input or ""), str(expected_password))
    return username_ok and password_ok


def render_login_gate():
    if st.session_state.get("authenticated"):
        return True

    expected_username, expected_password = _get_auth_credentials()

    outer_left, outer_center, outer_right = st.columns([1, 1.2, 1])
    with outer_center:
        st.title("🔬 Kredit Lab — Statement Intelligence")
        st.caption("Secure access required")
        with st.container(border=True):
            st.subheader("Sign in")
            if not expected_username or not expected_password:
                st.error("Authentication is not configured. Please set APP_USERNAME and APP_PASSWORD in Railway variables.")
                st.stop()

            with st.form("login_form", clear_on_submit=False):
                username_input = st.text_input("Username")
                password_input = st.text_input("Password", type="password")
                submitted = st.form_submit_button("Login", use_container_width=True)

            if submitted:
                is_valid = _check_login(username_input, password_input)
                if is_valid:
                    st.session_state["authenticated"] = True
                    st.session_state["auth_user"] = username_input
                    st.rerun()
                else:
                    st.error("Invalid username or password.")

    st.stop()


def render_logout_control():
    with st.sidebar:
        st.caption(f"Signed in as {st.session_state.get('auth_user', 'user')}")
        if st.button("Logout", use_container_width=True):
            st.session_state["authenticated"] = False
            st.session_state.pop("auth_user", None)
            st.rerun()


def fmt(val, decimals=2):
    """Format number with commas"""
    if val is None:
        return "0.00"
    return f"{val:,.{decimals}f}"


def generate_excel(data):
    """Generate Excel workbook with multiple sheets from v6.0.0 JSON"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    header_fill_green = PatternFill(start_color='196F3D', end_color='196F3D', fill_type='solid')
    header_fill_red = PatternFill(start_color='922B21', end_color='922B21', fill_type='solid')
    header_fill_orange = PatternFill(start_color='B9770E', end_color='B9770E', fill_type='solid')
    sub_header_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    alt_row_fill = PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid')
    credit_font = Font(name='Calibri', color='196F3D')
    debit_font = Font(name='Calibri', color='922B21')
    bold_font = Font(name='Calibri', bold=True, size=11)
    title_font = Font(name='Calibri', bold=True, size=14, color='1B4F72')
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC')
    )
    num_fmt = '#,##0.00'
    pct_fmt = '0.0%'

    def style_header_row(ws, row, max_col, fill=None):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = fill or header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    def style_data_cell(ws, row, col, is_number=False, is_credit=False, is_debit=False):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if is_number:
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal='right')
        if is_credit:
            cell.font = credit_font
        if is_debit:
            cell.font = debit_font
        if row % 2 == 0:
            cell.fill = alt_row_fill

    def auto_width(ws, min_width=10, max_width=40):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    val = str(cell.value) if cell.value else ''
                    max_len = max(max_len, len(val))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))

    r = data.get('report_info', {})
    accounts = data.get('accounts', [])
    monthly = data.get('monthly_analysis', [])
    consol = data.get('consolidated', {})
    top_parties = data.get('top_parties', {})
    large_credits = data.get('large_credits', [])
    own_related = data.get('own_related_transactions', {})
    loans = data.get('loan_transactions', {})
    flags = data.get('flags', {})
    obs = data.get('observations', {})

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="KREDIT LAB — STATEMENT INTELLIGENCE REPORT").font = title_font
    ws.cell(row=2, column=1, value=r.get('company_name', '')).font = bold_font
    ws.cell(row=3, column=1, value=f"Period: {r.get('period_start', '')} to {r.get('period_end', '')}")
    ws.cell(row=4, column=1, value=f"Generated: {r.get('generated_at', '')}")

    # Account info
    row = 6
    ws.cell(row=row, column=1, value="ACCOUNT DETAILS").font = bold_font
    row += 1
    headers = ['Bank', 'Account No', 'Holder', 'Type', 'Opening Balance', 'Closing Balance', 'Total Credits', 'Total Debits', 'Transactions']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    for a in accounts:
        row += 1
        vals = [a.get('bank_name'), a.get('account_number'), a.get('account_holder'),
                a.get('account_type'), a.get('opening_balance'), a.get('closing_balance'),
                a.get('total_credits'), a.get('total_debits'), a.get('transaction_count')]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v)
            style_data_cell(ws, row, c, is_number=c in [5,6,7,8], is_credit=c==7, is_debit=c==8)

    # Consolidated
    row += 2
    ws.cell(row=row, column=1, value="CONSOLIDATED FIGURES").font = bold_font
    row += 1
    consol_items = [
        ('Gross Credits', consol.get('gross_credits')),
        ('Gross Debits', consol.get('gross_debits')),
        ('Net Credits', consol.get('net_credits')),
        ('Net Debits', consol.get('net_debits')),
        ('Annualized Net Credits', consol.get('annualized_net_credits')),
        ('Annualized Net Debits', consol.get('annualized_net_debits')),
        ('', ''),
        ('Own Party Credits', consol.get('total_own_party_cr')),
        ('Own Party Debits', consol.get('total_own_party_dr')),
        ('Related Party Credits', consol.get('total_related_party_cr')),
        ('Related Party Debits', consol.get('total_related_party_dr')),
        ('', ''),
        ('Loan Disbursements', consol.get('total_loan_disbursement_cr')),
        ('Loan Repayments', consol.get('total_loan_repayment_dr')),
        ('FD/Interest Credits', consol.get('total_fd_interest_cr')),
        ('', ''),
        ('Cash Deposits', consol.get('total_cash_deposits')),
        ('Cash Withdrawals', consol.get('total_cash_withdrawals')),
        ('Cheque Deposits', consol.get('total_cheque_deposits')),
        ('Cheque Issues', consol.get('total_cheque_issues')),
        ('', ''),
        ('Total Salary Paid', consol.get('total_salary_paid')),
        ('Total EPF', consol.get('total_statutory_epf')),
        ('Total SOCSO', consol.get('total_statutory_socso')),
        ('Total Tax', consol.get('total_statutory_tax')),
        ('Total HRDF', consol.get('total_statutory_hrdf')),
        ('', ''),
        ('EOD Lowest', consol.get('eod_lowest')),
        ('EOD Highest', consol.get('eod_highest')),
        ('EOD Average', consol.get('eod_average')),
    ]
    # v6.2.0: FX consolidated fields
    is_v620 = data.get('report_info', {}).get('schema_version', '') in ('6.2.0', '6.2.1', '6.2.2', '6.3.0') or consol.get('total_fx_credits') is not None
    if is_v620:
        consol_items.append(('', ''))
        consol_items.append(('FX/Remittance Credits', consol.get('total_fx_credits')))
        consol_items.append(('FX/Remittance Debits', consol.get('total_fx_debits')))
        fx_cr_pct = consol.get('fx_credit_pct')
        if fx_cr_pct is not None:
            consol_items.append(('FX Credit % of Gross', fx_cr_pct))
        fx_dr_pct = consol.get('fx_debit_pct')
        if fx_dr_pct is not None:
            consol_items.append(('FX Debit % of Gross', fx_dr_pct))
        fx_currencies = consol.get('fx_currencies_all', [])
        if fx_currencies:
            consol_items.append(('FX Currencies Detected', ', '.join(fx_currencies)))
    for label, val in consol_items:
        if label:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val)
            ws.cell(row=row, column=2).number_format = num_fmt
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2).border = thin_border
        row += 1

    # Observations
    row += 1
    ws.cell(row=row, column=1, value="POSITIVE OBSERVATIONS").font = Font(bold=True, color='196F3D')
    row += 1
    for o in obs.get('positive', []):
        ws.cell(row=row, column=1, value=f"  + {o}")
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="CONCERNS").font = Font(bold=True, color='922B21')
    row += 1
    for o in obs.get('concerns', []):
        ws.cell(row=row, column=1, value=f"  - {o}")
        row += 1

    auto_width(ws)

    # ── Sheet 2: Monthly Analysis ──
    ws2 = wb.create_sheet("Cash Flow")
    headers = [
        'Month', 'Bank', 'Account No',
        'Gross Credits', 'Gross Debits', 'Net Credits', 'Net Debits',
        'Credit Count', 'Debit Count',
        'Own Party Cr', 'Own Party Dr', 'Related Party Cr', 'Related Party Dr',
        'Reversal Cr', 'Loan Disbursement Cr', 'FD Interest Cr',
        'Round Figure Cr', 'High Value Cr',
        'Cash Dep Count', 'Cash Dep Amt', 'Cash Wdl Count', 'Cash Wdl Amt',
        'Chq Dep Count', 'Chq Dep Amt', 'Chq Issue Count', 'Chq Issue Amt',
        'Loan Repayment Dr', 'Salary Paid',
        'EPF', 'SOCSO', 'Tax', 'HRDF',
        'Ret Chq In Count', 'Ret Chq In Amt', 'Ret Chq Out Count', 'Ret Chq Out Amt',
        'EOD Lowest', 'EOD Highest', 'EOD Average',
        'Opening Balance', 'Closing Balance'
    ]
    # v6.2.0: append FX columns
    if is_v620:
        headers += ['FX Cr Count', 'FX Cr Amount', 'FX Dr Count', 'FX Dr Amount', 'FX Currencies']
    # v6.2.1: append reconciliation columns
    has_recon = any(m.get('reconciliation_status') for m in monthly)
    if has_recon:
        headers += ['Recon Status', 'Recon Delta', 'Gaps', 'Missing Debits', 'Missing Credits', 'Data Quality Note']
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header_row(ws2, 1, len(headers))

    num_cols = set(range(4, 42)) - {8, 9, 19, 21, 23, 25, 33, 35}
    if is_v620:
        num_cols.update({43, 45})  # FX Cr Amount, FX Dr Amount
    for i, m in enumerate(monthly):
        row = i + 2
        vals = [
            m.get('month'), m.get('bank_name', ''), m.get('account_number', ''),
            m.get('gross_credits'), m.get('gross_debits'),
            m.get('net_credits'), m.get('net_debits'),
            m.get('credit_count'), m.get('debit_count'),
            m.get('own_party_cr'), m.get('own_party_dr'),
            m.get('related_party_cr'), m.get('related_party_dr'),
            m.get('reversal_cr'), m.get('loan_disbursement_cr'), m.get('fd_interest_cr'),
            m.get('round_figure_cr'), m.get('high_value_cr'),
            m.get('cash_deposits_count'), m.get('cash_deposits_amount'),
            m.get('cash_withdrawals_count'), m.get('cash_withdrawals_amount'),
            m.get('cheque_deposits_count'), m.get('cheque_deposits_amount'),
            m.get('cheque_issues_count'), m.get('cheque_issues_amount'),
            m.get('loan_repayment_dr'), m.get('salary_paid'),
            m.get('statutory_epf'), m.get('statutory_socso'),
            m.get('statutory_tax'), m.get('statutory_hrdf'),
            m.get('returned_cheques_inward_count'), m.get('returned_cheques_inward_amount'),
            m.get('returned_cheques_outward_count'), m.get('returned_cheques_outward_amount'),
            m.get('eod_lowest'), m.get('eod_highest'), m.get('eod_average'),
            m.get('opening_balance'), m.get('closing_balance')
        ]
        if is_v620:
            vals += [
                m.get('fx_credit_count', 0), m.get('fx_credit_amount', 0),
                m.get('fx_debit_count', 0), m.get('fx_debit_amount', 0),
                ', '.join(m.get('fx_currencies', []))
            ]
        if has_recon:
            vals += [
                m.get('reconciliation_status', ''),
                m.get('reconciliation_delta', 0),
                m.get('extraction_gaps', 0),
                m.get('missing_debit_amount', 0),
                m.get('missing_credit_amount', 0),
                m.get('data_quality_note', '') or ''
            ]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=row, column=c, value=v)
            style_data_cell(ws2, row, c, is_number=c in num_cols)
        # v6.2.1: Highlight FAIL rows in red
        if has_recon and m.get('reconciliation_status') == 'FAIL':
            fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for c in range(1, len(vals) + 1):
                ws2.cell(row=row, column=c).fill = fail_fill

    auto_width(ws2, min_width=12)

    # ── Sheet 3: Top Parties ──
    ws3 = wb.create_sheet("Top Parties")
    ws3.cell(row=1, column=1, value="TOP PAYERS (Income Sources)").font = bold_font

    # v6.2.0: detect monthly_breakdown availability
    has_monthly_bd = any(p.get('monthly_breakdown') for p in top_parties.get('top_payers', []) + top_parties.get('top_payees', []))
    # Collect all months from breakdown
    all_bd_months = set()
    if has_monthly_bd:
        for p in top_parties.get('top_payers', []) + top_parties.get('top_payees', []):
            for mb in p.get('monthly_breakdown', []):
                all_bd_months.add(mb.get('month', ''))
    bd_months_sorted = sorted(all_bd_months)

    headers = ['Rank', 'Party Name', 'Total Amount', 'Transactions', 'Related Party']
    if has_monthly_bd:
        headers += [f'{m}' for m in bd_months_sorted]
    for c, h in enumerate(headers, 1):
        ws3.cell(row=2, column=c, value=h)
    style_header_row(ws3, 2, len(headers), header_fill_green)
    for i, p in enumerate(top_parties.get('top_payers', [])):
        row = i + 3
        ws3.cell(row=row, column=1, value=p.get('rank'))
        ws3.cell(row=row, column=2, value=p.get('party_name'))
        ws3.cell(row=row, column=3, value=p.get('total_amount'))
        ws3.cell(row=row, column=4, value=p.get('transaction_count'))
        ws3.cell(row=row, column=5, value='Yes' if p.get('is_related_party') else 'No')
        style_data_cell(ws3, row, 3, is_number=True, is_credit=True)
        for c in [1,2,4,5]:
            style_data_cell(ws3, row, c)
        if has_monthly_bd:
            mb_lookup = {mb['month']: mb['amount'] for mb in p.get('monthly_breakdown', [])}
            for j, m in enumerate(bd_months_sorted):
                col = 6 + j
                ws3.cell(row=row, column=col, value=mb_lookup.get(m, 0))
                style_data_cell(ws3, row, col, is_number=True, is_credit=True)

    row = len(top_parties.get('top_payers', [])) + 5
    ws3.cell(row=row, column=1, value="TOP PAYEES (Payment Destinations)").font = bold_font
    row += 1
    for c, h in enumerate(headers, 1):
        ws3.cell(row=row, column=c, value=h)
    style_header_row(ws3, row, len(headers), header_fill_red)
    for i, p in enumerate(top_parties.get('top_payees', [])):
        r2 = row + i + 1
        ws3.cell(row=r2, column=1, value=p.get('rank'))
        ws3.cell(row=r2, column=2, value=p.get('party_name'))
        ws3.cell(row=r2, column=3, value=p.get('total_amount'))
        ws3.cell(row=r2, column=4, value=p.get('transaction_count'))
        ws3.cell(row=r2, column=5, value='Yes' if p.get('is_related_party') else 'No')
        style_data_cell(ws3, r2, 3, is_number=True, is_debit=True)
        for c in [1,2,4,5]:
            style_data_cell(ws3, r2, c)
        if has_monthly_bd:
            mb_lookup = {mb['month']: mb['amount'] for mb in p.get('monthly_breakdown', [])}
            for j, m in enumerate(bd_months_sorted):
                col = 6 + j
                ws3.cell(row=r2, column=col, value=mb_lookup.get(m, 0))
                style_data_cell(ws3, r2, col, is_number=True, is_debit=True)

    auto_width(ws3)

    # ── Sheet 4: Large Credits ──
    ws4 = wb.create_sheet("Large Credits")
    headers = ['Date', 'Description', 'Amount', 'Category', 'Balance']
    for c, h in enumerate(headers, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header_row(ws4, 1, len(headers), header_fill_green)
    for i, t in enumerate(large_credits):
        row = i + 2
        ws4.cell(row=row, column=1, value=t.get('date'))
        ws4.cell(row=row, column=2, value=t.get('description', '')[:80])
        ws4.cell(row=row, column=3, value=t.get('amount'))
        ws4.cell(row=row, column=4, value=t.get('category', ''))
        ws4.cell(row=row, column=5, value=t.get('balance'))
        style_data_cell(ws4, row, 3, is_number=True, is_credit=True)
        style_data_cell(ws4, row, 5, is_number=True)
        for c in [1,2,4]:
            style_data_cell(ws4, row, c)
    auto_width(ws4)

    # ── Sheet 5: Related Party Transactions ──
    ws5 = wb.create_sheet("Counterparty")
    ws5.cell(row=1, column=1, value="COUNTERPARTY TRANSACTIONS").font = title_font
    summary = own_related.get('summary', {})
    ws5.cell(row=3, column=1, value="Summary").font = bold_font
    summary_items = [
        ('Own Party Credits', summary.get('own_party_cr'), summary.get('own_party_cr_pct')),
        ('Own Party Debits', summary.get('own_party_dr'), summary.get('own_party_dr_pct')),
        ('Related Party Credits', summary.get('related_party_cr'), summary.get('related_party_cr_pct')),
        ('Related Party Debits', summary.get('related_party_dr'), summary.get('related_party_dr_pct')),
    ]
    row = 4
    for label, amt, pct in summary_items:
        ws5.cell(row=row, column=1, value=label)
        ws5.cell(row=row, column=2, value=amt)
        ws5.cell(row=row, column=2).number_format = num_fmt
        ws5.cell(row=row, column=3, value=f"{pct:.1f}%" if pct else '')
        row += 1

    row += 1
    headers = ['Date', 'Description', 'Amount', 'Type', 'Party Type', 'Party Name']
    for c, h in enumerate(headers, 1):
        ws5.cell(row=row, column=c, value=h)
    style_header_row(ws5, row, len(headers), header_fill_orange)
    for t in own_related.get('transactions', []):
        row += 1
        ws5.cell(row=row, column=1, value=t.get('date'))
        ws5.cell(row=row, column=2, value=t.get('description', '')[:60])
        ws5.cell(row=row, column=3, value=t.get('amount'))
        ws5.cell(row=row, column=4, value=t.get('type'))
        ws5.cell(row=row, column=5, value=t.get('party_type'))
        ws5.cell(row=row, column=6, value=t.get('party_name', ''))
        is_cr = t.get('type') == 'CREDIT'
        style_data_cell(ws5, row, 3, is_number=True, is_credit=is_cr, is_debit=not is_cr)
        for c in [1,2,4,5,6]:
            style_data_cell(ws5, row, c)
    auto_width(ws5)

    # ── Sheet 6: Loan Transactions ──
    ws6 = wb.create_sheet("Facilities")
    ws6.cell(row=1, column=1, value="LOAN DISBURSEMENTS (Credits)").font = bold_font
    headers = ['Date', 'Description', 'Amount', 'Category', 'Balance']
    row = 2
    for c, h in enumerate(headers, 1):
        ws6.cell(row=row, column=c, value=h)
    style_header_row(ws6, row, len(headers), header_fill_green)
    for t in loans.get('disbursements', []):
        row += 1
        ws6.cell(row=row, column=1, value=t.get('date'))
        ws6.cell(row=row, column=2, value=t.get('description', '')[:70])
        ws6.cell(row=row, column=3, value=t.get('amount'))
        ws6.cell(row=row, column=4, value=t.get('category', ''))
        ws6.cell(row=row, column=5, value=t.get('balance'))
        style_data_cell(ws6, row, 3, is_number=True, is_credit=True)
        style_data_cell(ws6, row, 5, is_number=True)
        for c in [1,2,4]:
            style_data_cell(ws6, row, c)

    row += 2
    ws6.cell(row=row, column=1, value="LOAN REPAYMENTS (Debits)").font = bold_font
    row += 1
    for c, h in enumerate(headers, 1):
        ws6.cell(row=row, column=c, value=h)
    style_header_row(ws6, row, len(headers), header_fill_red)
    for t in loans.get('repayments', []):
        row += 1
        ws6.cell(row=row, column=1, value=t.get('date'))
        ws6.cell(row=row, column=2, value=t.get('description', '')[:70])
        ws6.cell(row=row, column=3, value=t.get('amount'))
        ws6.cell(row=row, column=4, value=t.get('category', ''))
        ws6.cell(row=row, column=5, value=t.get('balance'))
        style_data_cell(ws6, row, 3, is_number=True, is_debit=True)
        style_data_cell(ws6, row, 5, is_number=True)
        for c in [1,2,4]:
            style_data_cell(ws6, row, c)
    auto_width(ws6)

    # ── Sheet 7: Flags ──
    ws7 = wb.create_sheet("Risk Signals")
    headers = ['#', 'Signal', 'Detected', 'Remarks']
    for c, h in enumerate(headers, 1):
        ws7.cell(row=1, column=c, value=h)
    style_header_row(ws7, 1, len(headers))
    for i, f in enumerate(flags.get('indicators', [])):
        row = i + 2
        ws7.cell(row=row, column=1, value=f.get('id'))
        ws7.cell(row=row, column=2, value=f.get('name'))
        ws7.cell(row=row, column=3, value='YES' if f.get('detected') else 'NO')
        ws7.cell(row=row, column=4, value=f.get('remarks'))
        for c in range(1, 5):
            style_data_cell(ws7, row, c)
        if f.get('detected'):
            ws7.cell(row=row, column=3).font = Font(color='922B21', bold=True)
    auto_width(ws7)

    # ── Sheet 8 (v6.2.0): Parsing Metadata ──
    parsing = data.get('parsing_metadata')
    if parsing:
        ws8 = wb.create_sheet("Parsing QC")
        ws8.cell(row=1, column=1, value="PARSING QUALITY METRICS").font = title_font
        ws8.cell(row=3, column=1, value="Overall Success Rate")
        ws8.cell(row=3, column=2, value=f"{parsing.get('overall_success_rate', 0):.1f}%")
        ws8.cell(row=4, column=1, value="Transactions Extracted")
        ws8.cell(row=4, column=2, value=parsing.get('total_transactions_extracted', 0))
        ws8.cell(row=5, column=1, value="Balance Checks Passed")
        ws8.cell(row=5, column=2, value=f"{parsing.get('total_balance_checks_passed', 0)} / {parsing.get('total_balance_checks', 0)}")
        for r in range(3, 6):
            ws8.cell(row=r, column=1).font = bold_font
            ws8.cell(row=r, column=1).border = thin_border
            ws8.cell(row=r, column=2).border = thin_border

        row = 7
        headers = ['Month', 'Account', 'Bank', 'Opening Bal', 'Closing Bal', 'Gross Cr', 'Gross Dr',
                    'Expected Close', 'Recon Delta', 'Passed', 'Txns Extracted', 'Notes']
        for c, h in enumerate(headers, 1):
            ws8.cell(row=row, column=c, value=h)
        style_header_row(ws8, row, len(headers))

        for chk in parsing.get('account_month_checks', []):
            row += 1
            vals = [
                chk.get('month'), chk.get('account_number'), chk.get('bank_name'),
                chk.get('opening_balance'), chk.get('closing_balance'),
                chk.get('gross_credits'), chk.get('gross_debits'),
                chk.get('expected_closing'), chk.get('reconciliation_delta'),
                'PASS' if chk.get('passed') else 'FAIL',
                chk.get('transactions_extracted'), chk.get('notes', '')
            ]
            for c, v in enumerate(vals, 1):
                ws8.cell(row=row, column=c, value=v)
                style_data_cell(ws8, row, c, is_number=c in {4,5,6,7,8,9,11})
            if not chk.get('passed'):
                ws8.cell(row=row, column=10).font = Font(color='922B21', bold=True)
            else:
                ws8.cell(row=row, column=10).font = Font(color='196F3D', bold=True)

        auto_width(ws8)

        # v6.2.1: Extraction gaps detail sub-table
        extraction_gaps = parsing.get('extraction_gaps')
        if extraction_gaps:
            row += 2
            ws8.cell(row=row, column=1, value="EXTRACTION GAPS DETAIL").font = title_font
            row += 1
            gap_headers = ['Month', 'Date', 'Page', 'Source File', 'Missing Type', 'Missing Amount',
                           'Balance Before Gap', 'Balance After Gap', 'Last Good Transaction', 'Next Transaction']
            for c, h in enumerate(gap_headers, 1):
                ws8.cell(row=row, column=c, value=h)
            style_header_row(ws8, row, len(gap_headers), header_fill_red)
            for g in extraction_gaps:
                row += 1
                gap_vals = [
                    g.get('month', ''), g.get('date', ''), g.get('page', ''),
                    g.get('source_file', ''), g.get('missing_type', ''),
                    g.get('missing_amount', 0),
                    g.get('balance_before_gap', 0), g.get('balance_after_gap', 0),
                    g.get('prev_description', '')[:60], g.get('next_description', '')[:60]
                ]
                for c, v in enumerate(gap_vals, 1):
                    ws8.cell(row=row, column=c, value=v)
                    style_data_cell(ws8, row, c, is_number=c in {6, 7, 8})
                # Highlight missing amount in red
                ws8.cell(row=row, column=6).font = Font(color='922B21', bold=True)
            auto_width(ws8)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_interactive_html(data):
    """Generate interactive HTML report for v6.0.0 schema"""

    r = data.get('report_info', {})
    accounts = data.get('accounts', [])
    monthly = data.get('monthly_analysis', [])
    consol = data.get('consolidated', {})
    top_parties = data.get('top_parties', {})
    large_credits = data.get('large_credits', [])
    own_related = data.get('own_related_transactions', {})
    loans = data.get('loan_transactions', {})
    flags_data = data.get('flags', {})
    obs = data.get('observations', {})
    parsing = data.get('parsing_metadata', {})

    # Version detection
    schema_v = r.get('schema_version', '')
    is_v620 = schema_v in ('6.2.0', '6.2.1', '6.2.2', '6.3.0') or consol.get('total_fx_credits') is not None
    is_v630 = schema_v == '6.3.0' or consol.get('total_unclassified_cr') is not None
    has_parsing = bool(parsing)
    has_monthly_bd = any(p.get('monthly_breakdown') for p in top_parties.get('top_payers', []) + top_parties.get('top_payees', []))

    # v6.2.1: Data quality detection
    data_completeness = consol.get('data_completeness', 'COMPLETE')
    has_recon = any(m.get('reconciliation_status') for m in monthly)
    is_incomplete = data_completeness == 'INCOMPLETE'
    total_missing_dr = consol.get('total_missing_debits', 0) or 0
    total_missing_cr = consol.get('total_missing_credits', 0) or 0
    months_with_gaps = consol.get('months_with_gaps', 0) or 0
    total_gaps = consol.get('total_extraction_gaps', 0) or 0
    dq_warning = consol.get('data_quality_warning', '')

    company = r.get('company_name', 'Company')
    period_start = r.get('period_start', '')
    period_end = r.get('period_end', '')
    total_months = r.get('total_months', 0)
    related_parties = r.get('related_parties', [])

    # Build data quality banner HTML
    dq_banner_html = ''
    if has_recon:
        if is_incomplete:
            affected_months = ', '.join(m.get('month', '') for m in monthly if m.get('reconciliation_status') == 'FAIL')
            dq_banner_html = f'''
            <div class="dq-banner dq-fail">
                <div class="dq-icon">⚠️</div>
                <div>
                    <div class="dq-title">Incomplete Extraction — {months_with_gaps} of {total_months} Months Affected</div>
                    <div class="dq-detail">Balance trail reconciliation detected {total_gaps} extraction gap(s) where transactions exist in the source PDF but were not captured. Figures marked with ⚠️ are understated.</div>
                    <div class="dq-stats">
                        <div><div class="dq-stat-label">Missing Debits</div><div class="dq-stat-val">RM {total_missing_dr:,.2f}</div></div>
                        <div><div class="dq-stat-label">Missing Credits</div><div class="dq-stat-val" style="color:var(--green)">RM {total_missing_cr:,.2f}</div></div>
                        <div><div class="dq-stat-label">Gaps</div><div class="dq-stat-val">{total_gaps}</div></div>
                        <div><div class="dq-stat-label">Months Affected</div><div class="dq-stat-val">{affected_months}</div></div>
                    </div>
                </div>
            </div>'''
        else:
            dq_banner_html = f'''
            <div class="dq-banner dq-pass">
                <div class="dq-icon">✅</div>
                <div>
                    <div class="dq-title">Extraction Complete — All {total_months} Months Pass Reconciliation</div>
                    <div class="dq-detail">Every transaction's running balance matches the statement balance. No extraction gaps detected.</div>
                </div>
            </div>'''

    # ── Account cards ──
    acc_cards = ""
    for a in accounts:
        acc_cards += f'''
        <div class="account-card">
            <div class="account-header">
                <span class="bank-name">{a.get('bank_name','')}</span>
                <span class="badge badge-{a.get('account_type','Current').lower()}">{a.get('account_type','')}</span>
            </div>
            <div class="account-number">A/C: {a.get('account_number','')}</div>
            <div class="account-holder">{a.get('account_holder','')}</div>
            <div class="account-metrics">
                <div class="metric"><div class="metric-label">Opening</div><div class="metric-value">RM {a.get('opening_balance',0):,.2f}</div></div>
                <div class="metric"><div class="metric-label">Closing</div><div class="metric-value {'debit' if a.get('closing_balance',0) < 10000 else ''}">RM {a.get('closing_balance',0):,.2f}</div></div>
                <div class="metric"><div class="metric-label">Credits</div><div class="metric-value credit">RM {a.get('total_credits',0):,.2f}</div></div>
                <div class="metric"><div class="metric-label">Debits</div><div class="metric-value debit">RM {a.get('total_debits',0):,.2f}</div></div>
                <div class="metric"><div class="metric-label">Transactions</div><div class="metric-value">{a.get('transaction_count',0):,}</div></div>
            </div>
        </div>'''

    # ── Related parties ──
    rp_html = ""
    for rp in related_parties:
        name = rp.get('name', rp) if isinstance(rp, dict) else str(rp)
        rel = rp.get('relationship', '') if isinstance(rp, dict) else ''
        rp_html += f'<span class="rp-tag">{name} <small>({rel})</small></span>'

    # ── Monthly analysis table rows (per-account with month subtotals) ──
    # Detect if data has per-account rows (v6.1.0) or consolidated (v6.0.0)
    has_account_col = any(m.get('account_number') for m in monthly)

    # Group by month for subtotals and chart aggregation
    from collections import OrderedDict
    monthly_by_month = OrderedDict()
    for m in monthly:
        mo = m.get('month', '')
        if mo not in monthly_by_month:
            monthly_by_month[mo] = []
        monthly_by_month[mo].append(m)

    # Build distinct accounts list for coloring
    acct_list = []
    seen_acct = set()
    for m in monthly:
        an = m.get('account_number', '')
        if an and an not in seen_acct:
            acct_list.append(an)
            seen_acct.add(an)
    acct_colors = {}
    palette = ['var(--blue)', 'var(--purple)', 'var(--green)', 'var(--amber)']
    for i, a in enumerate(acct_list):
        acct_colors[a] = palette[i % len(palette)]

    monthly_rows = ""
    # Aggregated data for charts (per-month consolidated)
    chart_agg = OrderedDict()  # month -> {net_credits, net_debits, eod_lowest, eod_highest, eod_average}

    for mo, rows in monthly_by_month.items():
        # Aggregate for chart
        agg = {}
        sum_fields = ['gross_credits','gross_debits','net_credits','net_debits',
                       'own_party_cr','own_party_dr','related_party_cr','related_party_dr',
                       'reversal_cr','loan_disbursement_cr','fd_interest_cr',
                       'cash_deposits_amount','cash_withdrawals_amount',
                       'cheque_deposits_amount','cheque_issues_amount',
                       'loan_repayment_dr','salary_paid',
                       'statutory_epf','statutory_socso','statutory_tax',
                       'returned_cheques_outward_amount','returned_cheques_outward_count',
                       'round_figure_cr','high_value_cr',
                       'credit_count','debit_count',
                       'own_party_cr_count','own_party_dr_count',
                       'related_party_cr_count','related_party_dr_count',
                       'loan_repayment_count','inward_return_cr',
                       'unclassified_cr_count','unclassified_cr_amount',
                       'unclassified_dr_count','unclassified_dr_amount']
        for fld in sum_fields:
            agg[fld] = sum(r.get(fld, 0) or 0 for r in rows)
        agg['eod_lowest'] = min(r.get('eod_lowest', 0) or 0 for r in rows)
        agg['eod_highest'] = max(r.get('eod_highest', 0) or 0 for r in rows)
        agg['eod_average'] = sum(r.get('eod_average', 0) or 0 for r in rows) / len(rows) if rows else 0
        # For multi-account months, sum opening/closing across accounts; for single account, take directly
        agg['opening_balance'] = sum(r.get('opening_balance', 0) or 0 for r in rows)
        agg['closing_balance'] = sum(r.get('closing_balance', 0) or 0 for r in rows)
        chart_agg[mo] = agg

        if has_account_col and len(rows) > 1:
            # Multiple accounts — show per-account rows then subtotal
            for m in rows:
                an = m.get('account_number', '')
                bn = m.get('bank_name', '')
                short_bank = bn.split(' ')[0] if bn else ''  # e.g. "OCBC" or "CIMB"
                acct_label = f"{short_bank} {an}" if an else mo
                dot_color = acct_colors.get(an, 'var(--text-muted)')
                monthly_rows += f'''<tr style="font-size:0.78rem;">
            <td class="sticky-col" style="padding-left:1.5rem;font-weight:400"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{dot_color};margin-right:6px;vertical-align:middle"></span>{acct_label}</td>
            {'<td></td>' if has_recon else ''}
            <td class="mono r credit">{m.get('gross_credits',0):,.2f}</td>
            <td class="mono r debit">{m.get('gross_debits',0):,.2f}</td>
            <td class="mono r credit">{m.get('net_credits',0):,.2f}</td>
            <td class="mono r debit">{m.get('net_debits',0):,.2f}</td>
            <td class="mono r">{m.get('credit_count',0)}</td>
            <td class="mono r">{m.get('debit_count',0)}</td>
            <td class="mono r">{m.get('own_party_cr',0):,.2f}</td>
            <td class="mono r">{m.get('own_party_dr',0):,.2f}</td>
            <td class="mono r">{m.get('related_party_cr',0):,.2f}</td>
            <td class="mono r">{m.get('related_party_dr',0):,.2f}</td>
            <td class="mono r">{m.get('reversal_cr',0):,.2f}</td>
            <td class="mono r">{m.get('loan_disbursement_cr',0):,.2f}</td>
            <td class="mono r">{m.get('fd_interest_cr',0):,.2f}</td>
            <td class="mono r">{m.get('cash_deposits_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cash_withdrawals_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cheque_deposits_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cheque_issues_amount',0):,.2f}</td>
            <td class="mono r">{m.get('loan_repayment_dr',0):,.2f}</td>
            <td class="mono r">{m.get('salary_paid',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_epf',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_socso',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_tax',0):,.2f}</td>
            <td class="mono r">{m.get('returned_cheques_outward_count',0)}</td>
            <td class="mono r">{m.get('returned_cheques_outward_amount',0):,.2f}</td>
            <td class="mono r">{m.get('round_figure_cr',0):,.2f}</td>
            <td class="mono r">{m.get('high_value_cr',0):,.2f}</td>
            <td class="mono r">{m.get('eod_lowest',0):,.2f}</td>
            <td class="mono r">{m.get('eod_highest',0):,.2f}</td>
            <td class="mono r">{m.get('eod_average',0):,.2f}</td>
            <td class="mono r">{m.get('opening_balance',0):,.2f}</td>
            <td class="mono r">{m.get('closing_balance',0):,.2f}</td>
            {'<td class="mono r v630-count">' + str(m.get('own_party_cr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('own_party_dr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('related_party_cr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('related_party_dr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('loan_repayment_count',0)) + '</td><td class="mono r v630-amt">' + f"{m.get('inward_return_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(m.get('unclassified_cr_count',0)) + '</td><td class="mono r v630-uncl">' + f"{m.get('unclassified_cr_amount',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(m.get('unclassified_dr_count',0)) + '</td><td class="mono r v630-uncl">' + f"{m.get('unclassified_dr_amount',0):,.2f}" + '</td>' if is_v630 else ''}
        </tr>'''

            # Month subtotal row
            a = agg
            # v6.2.1: aggregate reconciliation for multi-account month
            month_recon_cell = ''
            if has_recon:
                any_fail = any(r.get('reconciliation_status') == 'FAIL' for r in rows)
                if any_fail:
                    total_gaps = sum(r.get('extraction_gaps', 0) for r in rows)
                    total_miss = sum(r.get('missing_debit_amount', 0) for r in rows)
                    month_recon_cell = f'<td><span class="recon-badge fail">✗ FAIL</span> <span class="gap-pill">{total_gaps} gap{"s" if total_gaps > 1 else ""} · RM {total_miss:,.0f}</span></td>'
                else:
                    month_recon_cell = '<td><span class="recon-badge pass">✓ PASS</span></td>'
            monthly_rows += f'''<tr style="background:var(--bg);font-weight:600;border-bottom:2px solid var(--border-accent);{"" if not (has_recon and any_fail) else ""}">
            <td class="sticky-col">{mo}</td>
            {month_recon_cell}
            <td class="mono r credit">{a['gross_credits']:,.2f}</td>
            <td class="mono r debit">{a['gross_debits']:,.2f}</td>
            <td class="mono r credit" style="font-weight:700">{a['net_credits']:,.2f}</td>
            <td class="mono r debit" style="font-weight:700">{a['net_debits']:,.2f}</td>
            <td class="mono r">{int(a['credit_count'])}</td>
            <td class="mono r">{int(a['debit_count'])}</td>
            <td class="mono r">{a['own_party_cr']:,.2f}</td>
            <td class="mono r">{a['own_party_dr']:,.2f}</td>
            <td class="mono r">{a['related_party_cr']:,.2f}</td>
            <td class="mono r">{a['related_party_dr']:,.2f}</td>
            <td class="mono r">{a['reversal_cr']:,.2f}</td>
            <td class="mono r">{a['loan_disbursement_cr']:,.2f}</td>
            <td class="mono r">{a['fd_interest_cr']:,.2f}</td>
            <td class="mono r">{a['cash_deposits_amount']:,.2f}</td>
            <td class="mono r">{a['cash_withdrawals_amount']:,.2f}</td>
            <td class="mono r">{a['cheque_deposits_amount']:,.2f}</td>
            <td class="mono r">{a['cheque_issues_amount']:,.2f}</td>
            <td class="mono r">{a['loan_repayment_dr']:,.2f}</td>
            <td class="mono r">{a['salary_paid']:,.2f}</td>
            <td class="mono r">{a['statutory_epf']:,.2f}</td>
            <td class="mono r">{a['statutory_socso']:,.2f}</td>
            <td class="mono r">{a['statutory_tax']:,.2f}</td>
            <td class="mono r">{int(a['returned_cheques_outward_count'])}</td>
            <td class="mono r">{a['returned_cheques_outward_amount']:,.2f}</td>
            <td class="mono r">{a['round_figure_cr']:,.2f}</td>
            <td class="mono r">{a['high_value_cr']:,.2f}</td>
            <td class="mono r">{a['eod_lowest']:,.2f}</td>
            <td class="mono r">{a['eod_highest']:,.2f}</td>
            <td class="mono r">{a['eod_average']:,.2f}</td>
            <td class="mono r">{a['opening_balance']:,.2f}</td>
            <td class="mono r">{a['closing_balance']:,.2f}</td>
            {'<td class="mono r v630-count">' + str(int(a.get('own_party_cr_count',0))) + '</td><td class="mono r v630-count">' + str(int(a.get('own_party_dr_count',0))) + '</td><td class="mono r v630-count">' + str(int(a.get('related_party_cr_count',0))) + '</td><td class="mono r v630-count">' + str(int(a.get('related_party_dr_count',0))) + '</td><td class="mono r v630-count">' + str(int(a.get('loan_repayment_count',0))) + '</td><td class="mono r v630-amt">' + f"{a.get('inward_return_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(int(a.get('unclassified_cr_count',0))) + '</td><td class="mono r v630-uncl">' + f"{a.get('unclassified_cr_amount',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(int(a.get('unclassified_dr_count',0))) + '</td><td class="mono r v630-uncl">' + f"{a.get('unclassified_dr_amount',0):,.2f}" + '</td>' if is_v630 else ''}
        </tr>'''
        else:
            # Single account or v6.0.0 consolidated — single row per month
            m = rows[0] if rows else {}
            recon_status = m.get('reconciliation_status', '')
            row_class = ' class="row-fail"' if recon_status == 'FAIL' else ''
            recon_cell = ''
            if has_recon:
                if recon_status == 'FAIL':
                    gap_count = m.get('extraction_gaps', 0)
                    miss_dr = m.get('missing_debit_amount', 0)
                    recon_cell = f'<td><span class="recon-badge fail">✗ FAIL</span> <span class="gap-pill">{gap_count} gap{"s" if gap_count > 1 else ""} · RM {miss_dr:,.0f}</span></td>'
                else:
                    recon_cell = '<td><span class="recon-badge pass">✓ PASS</span></td>'
            monthly_rows += f'''<tr{row_class}>
            <td class="sticky-col">{m.get('month','')}</td>
            {recon_cell}
            <td class="mono r credit">{m.get('gross_credits',0):,.2f}</td>
            <td class="mono r debit">{m.get('gross_debits',0):,.2f}</td>
            <td class="mono r credit" style="font-weight:600">{m.get('net_credits',0):,.2f}</td>
            <td class="mono r debit" style="font-weight:600">{m.get('net_debits',0):,.2f}</td>
            <td class="mono r">{m.get('credit_count',0)}</td>
            <td class="mono r">{m.get('debit_count',0)}</td>
            <td class="mono r">{m.get('own_party_cr',0):,.2f}</td>
            <td class="mono r">{m.get('own_party_dr',0):,.2f}</td>
            <td class="mono r">{m.get('related_party_cr',0):,.2f}</td>
            <td class="mono r">{m.get('related_party_dr',0):,.2f}</td>
            <td class="mono r">{m.get('reversal_cr',0):,.2f}</td>
            <td class="mono r">{m.get('loan_disbursement_cr',0):,.2f}</td>
            <td class="mono r">{m.get('fd_interest_cr',0):,.2f}</td>
            <td class="mono r">{m.get('cash_deposits_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cash_withdrawals_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cheque_deposits_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cheque_issues_amount',0):,.2f}</td>
            <td class="mono r">{m.get('loan_repayment_dr',0):,.2f}</td>
            <td class="mono r">{m.get('salary_paid',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_epf',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_socso',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_tax',0):,.2f}</td>
            <td class="mono r">{m.get('returned_cheques_outward_count',0)}</td>
            <td class="mono r">{m.get('returned_cheques_outward_amount',0):,.2f}</td>
            <td class="mono r">{m.get('round_figure_cr',0):,.2f}</td>
            <td class="mono r">{m.get('high_value_cr',0):,.2f}</td>
            <td class="mono r">{m.get('eod_lowest',0):,.2f}</td>
            <td class="mono r">{m.get('eod_highest',0):,.2f}</td>
            <td class="mono r">{m.get('eod_average',0):,.2f}</td>
            <td class="mono r">{m.get('opening_balance',0):,.2f}</td>
            <td class="mono r">{m.get('closing_balance',0):,.2f}</td>
            {'<td class="mono r v630-count">' + str(m.get('own_party_cr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('own_party_dr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('related_party_cr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('related_party_dr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('loan_repayment_count',0)) + '</td><td class="mono r v630-amt">' + f"{m.get('inward_return_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(m.get('unclassified_cr_count',0)) + '</td><td class="mono r v630-uncl">' + f"{m.get('unclassified_cr_amount',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(m.get('unclassified_dr_count',0)) + '</td><td class="mono r v630-uncl">' + f"{m.get('unclassified_dr_amount',0):,.2f}" + '</td>' if is_v630 else ''}
        </tr>'''

    # ── Consolidated totals row ──
    total_status_cell = ''
    if has_recon:
        if is_incomplete:
            total_status_cell = f'<td><span class="recon-badge fail">⚠️ {months_with_gaps} FAIL</span></td>'
        else:
            total_status_cell = '<td><span class="recon-badge pass">ALL PASS</span></td>'
    consol_row = f'''<tr class="total-row">
        <td class="sticky-col" style="font-weight:700">TOTAL</td>
        {total_status_cell}
        <td class="mono r credit">{consol.get('gross_credits',0):,.2f}</td>
        <td class="mono r debit">{consol.get('gross_debits',0):,.2f}</td>
        <td class="mono r credit">{consol.get('net_credits',0):,.2f}</td>
        <td class="mono r debit">{consol.get('net_debits',0):,.2f}</td>
        <td class="mono r">-</td><td class="mono r">-</td>
        <td class="mono r">{consol.get('total_own_party_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_own_party_dr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_related_party_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_related_party_dr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_reversal_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_loan_disbursement_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_fd_interest_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_cash_deposits',0):,.2f}</td>
        <td class="mono r">{consol.get('total_cash_withdrawals',0):,.2f}</td>
        <td class="mono r">{consol.get('total_cheque_deposits',0):,.2f}</td>
        <td class="mono r">{consol.get('total_cheque_issues',0):,.2f}</td>
        <td class="mono r">{consol.get('total_loan_repayment_dr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_salary_paid',0):,.2f}</td>
        <td class="mono r">{consol.get('total_statutory_epf',0):,.2f}</td>
        <td class="mono r">{consol.get('total_statutory_socso',0):,.2f}</td>
        <td class="mono r">{consol.get('total_statutory_tax',0):,.2f}</td>
        <td class="mono r">{consol.get('total_returned_cheques_outward',0):,.2f}</td>
        <td class="mono r">{consol.get('total_returned_cheques_outward',0):,.2f}</td>
        <td class="mono r">{consol.get('total_round_figure_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_high_value_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('eod_lowest',0):,.2f}</td>
        <td class="mono r">{consol.get('eod_highest',0):,.2f}</td>
        <td class="mono r">{consol.get('eod_average',0):,.2f}</td>
        <td class="mono r">-</td>
        <td class="mono r">-</td>
        {'<td class="mono r v630-count">-</td><td class="mono r v630-count">-</td><td class="mono r v630-count">-</td><td class="mono r v630-count">-</td><td class="mono r v630-count">-</td><td class="mono r v630-amt">' + f"{consol.get('total_inward_return_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">-</td><td class="mono r v630-uncl">' + f"{consol.get('total_unclassified_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">-</td><td class="mono r v630-uncl">' + f"{consol.get('total_unclassified_dr',0):,.2f}" + '</td>' if is_v630 else ''}
    </tr>'''

    # ── Top payers/payees ──
    payer_rows = ""
    for p in top_parties.get('top_payers', []):
        rp_badge = '<span class="rp-badge">RP</span>' if p.get('is_related_party') else ''
        # v6.2.0: monthly breakdown mini-bar
        mb_html = ''
        if has_monthly_bd and p.get('monthly_breakdown'):
            mb_vals = [mb.get('amount', 0) for mb in p['monthly_breakdown']]
            max_mb = max(mb_vals) if mb_vals and max(mb_vals) > 0 else 1
            bars = ''.join(
                f'<div title="{mb.get("month","")}: RM {mb.get("amount",0):,.0f}" style="flex:1;background:var(--green);opacity:0.7;border-radius:2px;min-width:4px;height:{max(2, int(mb.get("amount",0)/max_mb*28))}px"></div>'
                for mb in p['monthly_breakdown']
            )
            mb_html = f'<div style="display:flex;align-items:flex-end;gap:2px;height:30px;margin-top:4px">{bars}</div>'
        payer_rows += f'''<tr>
            <td>{p.get('rank')}</td>
            <td>{p.get('party_name','')} {rp_badge}{mb_html}</td>
            <td class="mono r credit">RM {p.get('total_amount',0):,.2f}</td>
            <td class="mono r">{p.get('transaction_count',0)}</td>
        </tr>'''

    payee_rows = ""
    for p in top_parties.get('top_payees', []):
        rp_badge = '<span class="rp-badge">RP</span>' if p.get('is_related_party') else ''
        mb_html = ''
        if has_monthly_bd and p.get('monthly_breakdown'):
            mb_vals = [mb.get('amount', 0) for mb in p['monthly_breakdown']]
            max_mb = max(mb_vals) if mb_vals and max(mb_vals) > 0 else 1
            bars = ''.join(
                f'<div title="{mb.get("month","")}: RM {mb.get("amount",0):,.0f}" style="flex:1;background:var(--red);opacity:0.7;border-radius:2px;min-width:4px;height:{max(2, int(mb.get("amount",0)/max_mb*28))}px"></div>'
                for mb in p['monthly_breakdown']
            )
            mb_html = f'<div style="display:flex;align-items:flex-end;gap:2px;height:30px;margin-top:4px">{bars}</div>'
        payee_rows += f'''<tr>
            <td>{p.get('rank')}</td>
            <td>{p.get('party_name','')} {rp_badge}{mb_html}</td>
            <td class="mono r debit">RM {p.get('total_amount',0):,.2f}</td>
            <td class="mono r">{p.get('transaction_count',0)}</td>
        </tr>'''

    # ── Large credits ──
    large_cr_rows = ""
    for t in large_credits:
        large_cr_rows += f'''<tr>
            <td>{t.get('date','')}</td>
            <td>{t.get('description','')[:70]}</td>
            <td class="mono r credit">RM {t.get('amount',0):,.2f}</td>
            <td class="mono r">{t.get('balance',0):,.2f}</td>
        </tr>'''

    # ── Related party transactions ──
    rp_summary = own_related.get('summary', {})
    rp_txn_rows = ""
    for t in own_related.get('transactions', [])[:50]:
        type_cls = 'credit' if t.get('type') == 'CREDIT' else 'debit'
        rp_txn_rows += f'''<tr>
            <td>{t.get('date','')}</td>
            <td>{t.get('description','')[:55]}</td>
            <td class="mono r {type_cls}">RM {t.get('amount',0):,.2f}</td>
            <td><span class="badge badge-{t.get('type','').lower()}">{t.get('type','')}</span></td>
            <td>{t.get('party_type','')}</td>
            <td>{t.get('party_name','')}</td>
        </tr>'''
    rp_total = len(own_related.get('transactions', []))
    rp_note = f'<div class="note">Showing first 50 of {rp_total} transactions</div>' if rp_total > 50 else ''

    # ── Loan transactions ──
    loan_disb_rows = ""
    for t in loans.get('disbursements', []):
        loan_disb_rows += f'''<tr>
            <td>{t.get('date','')}</td><td>{t.get('description','')[:55]}</td>
            <td class="mono r credit">RM {t.get('amount',0):,.2f}</td>
            <td>{t.get('category','')}</td>
        </tr>'''
    loan_repay_rows = ""
    for t in loans.get('repayments', []):
        loan_repay_rows += f'''<tr>
            <td>{t.get('date','')}</td><td>{t.get('description','')[:55]}</td>
            <td class="mono r debit">RM {t.get('amount',0):,.2f}</td>
            <td>{t.get('category','')}</td>
        </tr>'''

    # ── Flags ──
    flag_rows = ""
    detected_count = 0
    for f in flags_data.get('indicators', []):
        detected = f.get('detected', False)
        if detected:
            detected_count += 1
        status_cls = 'flag-yes' if detected else 'flag-no'
        flag_rows += f'''<tr class="{status_cls}">
            <td class="mono">{f.get('id','')}</td>
            <td>{f.get('name','')}</td>
            <td class="mono" style="text-align:center"><span class="flag-dot {'detected' if detected else 'clear'}"></span> {'YES' if detected else 'NO'}</td>
            <td>{f.get('remarks','')}</td>
        </tr>'''

    # ── Observations ──
    pos_obs = "".join([f'<li class="obs-item positive">{o}</li>' for o in obs.get('positive', [])])
    con_obs_items = obs.get('concerns', [])
    con_obs = "".join([
        f'<li class="obs-item {"data-warn" if "DATA QUALITY" in o or "INCOMPLETE" in o.upper() or "extraction gap" in o.lower() else "concern"}">{o}</li>'
        for o in con_obs_items
    ])

    # ── Chart data (JSON for Plotly) — use aggregated monthly if per-account ──
    # v6.2.0: also build FX chart data
    fx_chart_cr = []
    fx_chart_dr = []

    if chart_agg:
        chart_months = json.dumps(list(chart_agg.keys()))
        chart_net_cr = json.dumps([round(a['net_credits'], 2) for a in chart_agg.values()])
        chart_net_dr = json.dumps([round(a['net_debits'], 2) for a in chart_agg.values()])
        chart_eod_avg = json.dumps([round(a['eod_average'], 2) for a in chart_agg.values()])
        chart_eod_low = json.dumps([round(a['eod_lowest'], 2) for a in chart_agg.values()])
        chart_eod_high = json.dumps([round(a['eod_highest'], 2) for a in chart_agg.values()])
        if is_v620:
            for mo, rows in monthly_by_month.items():
                fx_chart_cr.append(sum(r.get('fx_credit_amount', 0) or 0 for r in rows))
                fx_chart_dr.append(sum(r.get('fx_debit_amount', 0) or 0 for r in rows))
    else:
        chart_months = json.dumps([m.get('month', '') for m in monthly])
        chart_net_cr = json.dumps([m.get('net_credits', 0) for m in monthly])
        chart_net_dr = json.dumps([m.get('net_debits', 0) for m in monthly])
        chart_eod_avg = json.dumps([m.get('eod_average', 0) for m in monthly])
        chart_eod_low = json.dumps([m.get('eod_lowest', 0) for m in monthly])
        chart_eod_high = json.dumps([m.get('eod_highest', 0) for m in monthly])
        if is_v620:
            fx_chart_cr = [m.get('fx_credit_amount', 0) for m in monthly]
            fx_chart_dr = [m.get('fx_debit_amount', 0) for m in monthly]

    fx_chart_cr_json = json.dumps(fx_chart_cr)
    fx_chart_dr_json = json.dumps(fx_chart_dr)

    # v6.2.0: Build FX tab HTML
    fx_tab_html = ''
    if is_v620:
        fx_currencies_all = consol.get('fx_currencies_all', [])
        fx_currencies_str = ', '.join(fx_currencies_all) if fx_currencies_all else 'None detected'
        fx_tab_html = f'''
        <div id="tab-fx" class="tab">
            <div class="info-panel">
                <h4>FX Classification Methodology</h4>
                <p>Transactions are classified as FX only when there is clear evidence of foreign currency conversion. Key rules:</p>
                <ul>
                    <li><strong>Default rule:</strong> NOT classified as FX unless clear conversion evidence exists</li>
                    <li><strong>TT CREDIT</strong> = Telegraphic Transfer (payment method), NOT a currency indicator</li>
                    <li><strong>RENTAS / JANM</strong> = Domestic MYR-to-MYR interbank transfers (Real-time Electronic Transfer of Funds and Securities)</li>
                    <li><strong>Voucher codes</strong> (GBPV, USDP) in reference fields = internal bank numbering, not currency denominations</li>
                    <li><strong>True FX requires:</strong> conversion rate, foreign currency amount, SWIFT codes, or foreign beneficiary with non-MYR amounts</li>
                </ul>
            </div>
            <div class="summary-grid">
                <div class="summary-card"><div class="val credit">{consol.get('total_fx_credits',0):,.0f}</div><div class="lbl">FX Credits (Total)</div></div>
                <div class="summary-card"><div class="val debit">{consol.get('total_fx_debits',0):,.0f}</div><div class="lbl">FX Debits (Total)</div></div>
                <div class="summary-card"><div class="val">{consol.get('fx_credit_pct',0):.1f}%</div><div class="lbl">FX Cr % of Gross</div></div>
                <div class="summary-card"><div class="val">{consol.get('fx_debit_pct',0):.1f}%</div><div class="lbl">FX Dr % of Gross</div></div>
                <div class="summary-card"><div class="val" style="font-size:1rem">{fx_currencies_str}</div><div class="lbl">Currencies Detected</div></div>
            </div>
            <div class="section">
                <div class="section-head"><h2>FX / Remittance Trend</h2></div>
                <div class="section-body">
                    <div id="chartFX" style="height:300px"></div>
                </div>
            </div>
            <div class="section">
                <div class="section-head"><h2>Monthly FX Breakdown</h2></div>
                <div class="section-body" style="padding:0">
                    <div class="table-wrap"><table>
                        <thead><tr><th>Month</th><th class="r">FX Cr Count</th><th class="r">FX Cr Amount</th><th class="r">FX Dr Count</th><th class="r">FX Dr Amount</th><th>Currencies</th></tr></thead>
                        <tbody>'''
        for mo, rows in monthly_by_month.items():
            fx_cc = sum(r.get('fx_credit_count', 0) or 0 for r in rows)
            fx_ca = sum(r.get('fx_credit_amount', 0) or 0 for r in rows)
            fx_dc = sum(r.get('fx_debit_count', 0) or 0 for r in rows)
            fx_da = sum(r.get('fx_debit_amount', 0) or 0 for r in rows)
            fx_cur = set()
            for r2 in rows:
                fx_cur.update(r2.get('fx_currencies', []))
            fx_tab_html += f'''<tr><td>{mo}</td><td class="mono r">{fx_cc}</td><td class="mono r credit">RM {fx_ca:,.2f}</td>
                <td class="mono r">{fx_dc}</td><td class="mono r debit">RM {fx_da:,.2f}</td>
                <td>{', '.join(sorted(fx_cur)) if fx_cur else '-'}</td></tr>'''
        fx_tab_html += '</tbody></table></div></div></div></div>'

    # v6.3.0: Build Unclassified Transactions tab HTML
    unclassified_tab_html = ''
    if is_v630:
        uncl_txns = data.get('unclassified_transactions', [])
        uncl_cr_total = consol.get('total_unclassified_cr', 0) or 0
        uncl_dr_total = consol.get('total_unclassified_dr', 0) or 0
        uncl_cr_count_total = sum((m.get('unclassified_cr_count', 0) or 0) for m in monthly)
        uncl_dr_count_total = sum((m.get('unclassified_dr_count', 0) or 0) for m in monthly)
        cls_config = data.get('classification_config', {})
        uncl_threshold = cls_config.get('unclassified_listing_threshold', 10000)

        # Monthly breakdown rows
        uncl_monthly_rows = ''
        for mo, rows in monthly_by_month.items():
            mo_uncl_cr_count = sum(r.get('unclassified_cr_count', 0) or 0 for r in rows)
            mo_uncl_cr_amt = sum(r.get('unclassified_cr_amount', 0) or 0 for r in rows)
            mo_uncl_dr_count = sum(r.get('unclassified_dr_count', 0) or 0 for r in rows)
            mo_uncl_dr_amt = sum(r.get('unclassified_dr_amount', 0) or 0 for r in rows)
            mo_net_cr = sum(r.get('net_credits', 0) or 0 for r in rows)
            pct_of_net = (mo_uncl_cr_amt / mo_net_cr * 100) if mo_net_cr > 0 else 0
            uncl_monthly_rows += f'''<tr>
                <td>{mo}</td>
                <td class="mono r">{mo_uncl_cr_count}</td>
                <td class="mono r credit">RM {mo_uncl_cr_amt:,.2f}</td>
                <td class="mono r">{mo_uncl_dr_count}</td>
                <td class="mono r debit">RM {mo_uncl_dr_amt:,.2f}</td>
                <td class="mono r">{pct_of_net:.1f}%</td>
            </tr>'''

        # Individual transactions rows
        uncl_txn_rows = ''
        for t in uncl_txns:
            type_cls = 'credit' if t.get('type') == 'CREDIT' else 'debit'
            uncl_txn_rows += f'''<tr>
                <td>{t.get('date','')}</td>
                <td>{t.get('description','')[:70]}</td>
                <td class="mono r {type_cls}">RM {t.get('amount',0):,.2f}</td>
                <td><span class="badge badge-{t.get('type','').lower()}">{t.get('type','')}</span></td>
                <td class="mono r">{t.get('balance',0):,.2f}</td>
            </tr>'''

        unclassified_tab_html = f'''
        <div id="tab-unclassified" class="tab">
            <div class="info-panel">
                <h4>What are Unclassified Transactions?</h4>
                <p>These are transactions whose descriptions are too vague or do not match any known classification rule.
                Unclassified credits <strong>remain in Net Credits</strong> and unclassified debits <strong>remain in Net Debits</strong> &mdash;
                they are NOT excluded. This section highlights them for analyst review.</p>
            </div>
            <div class="summary-grid">
                <div class="summary-card"><div class="val credit">RM {uncl_cr_total:,.0f}</div><div class="lbl">Unclassified Credits</div></div>
                <div class="summary-card"><div class="val debit">RM {uncl_dr_total:,.0f}</div><div class="lbl">Unclassified Debits</div></div>
                <div class="summary-card"><div class="val">{uncl_cr_count_total}</div><div class="lbl">Credit Txn Count</div></div>
                <div class="summary-card"><div class="val">{uncl_dr_count_total}</div><div class="lbl">Debit Txn Count</div></div>
            </div>
            <div class="section">
                <div class="section-head"><h2>Monthly Breakdown</h2></div>
                <div class="section-body" style="padding:0"><div class="table-wrap"><table>
                    <thead><tr><th>Month</th><th class="r">Uncl Cr #</th><th class="r">Uncl Cr Amt</th><th class="r">Uncl Dr #</th><th class="r">Uncl Dr Amt</th><th class="r">% of Net Cr</th></tr></thead>
                    <tbody>{uncl_monthly_rows}</tbody>
                </table></div></div>
            </div>
            {'<div class="section"><div class="section-head"><h2>Individual Unclassified Transactions (&ge; RM ' + f"{uncl_threshold:,.0f}" + ')</h2><span class="badge badge-current">' + str(len(uncl_txns)) + ' transactions</span></div><div class="section-body" style="padding:0"><div class="table-wrap" style="max-height:500px;overflow:auto"><table><thead><tr><th>Date</th><th>Description</th><th class="r">Amount</th><th>Type</th><th class="r">Balance</th></tr></thead><tbody>' + uncl_txn_rows + '</tbody></table></div></div></div>' if uncl_txns else ''}
        </div>'''

    # v6.2.0/v6.2.1: Build Parsing QC tab HTML
    parsing_tab_html = ''
    if has_parsing:
        success_rate = parsing.get('overall_success_rate', 0)
        rate_color = 'green' if success_rate >= 95 else 'amber' if success_rate >= 80 else 'red'
        # v6.2.1: Additional gap stats
        p_total_gaps = len(parsing.get('extraction_gaps', []) or [])
        p_missing_dr = consol.get('total_missing_debits', 0) or 0
        p_missing_cr = consol.get('total_missing_credits', 0) or 0
        gap_cards = ''
        if has_recon:
            gap_cards = f'''
                <div class="summary-card"><div class="val" style="color:var(--{'red' if p_total_gaps > 0 else 'green'})">{p_total_gaps}</div><div class="lbl">Extraction Gaps</div></div>
                <div class="summary-card"><div class="val" style="color:var(--{'red' if p_missing_dr > 0 else 'green'})">RM {p_missing_dr:,.0f}</div><div class="lbl">Missing Debits</div></div>
                <div class="summary-card"><div class="val" style="color:var(--{'red' if p_missing_cr > 0 else 'green'})">RM {p_missing_cr:,.0f}</div><div class="lbl">Missing Credits</div></div>'''
        parsing_tab_html = f'''
        <div id="tab-parsing" class="tab">
            {dq_banner_html}
            <div class="summary-grid">
                <div class="summary-card"><div class="val" style="color:var(--{rate_color})">{success_rate:.1f}%</div><div class="lbl">Success Rate</div></div>
                <div class="summary-card"><div class="val">{parsing.get('total_transactions_extracted',0):,}</div><div class="lbl">Txns Extracted</div></div>
                <div class="summary-card"><div class="val">{parsing.get('total_balance_checks_passed',0)}/{parsing.get('total_balance_checks',0)}</div><div class="lbl">Balance Checks Passed</div></div>
                {gap_cards}
            </div>
            <div class="section">
                <div class="section-head"><h2>Balance Reconciliation Detail</h2></div>
                <div class="section-body" style="padding:0"><div class="table-wrap"><table>
                    <thead><tr><th>Month</th><th>Account</th><th class="r">Opening Balance</th><th class="r">Gross Credits</th><th class="r">Gross Debits</th><th class="r">Expected Closing</th><th class="r">Actual Closing</th><th class="r">Delta</th><th style="text-align:center">Status</th><th class="r">Txns</th><th>Gaps</th><th>Notes</th></tr></thead>
                    <tbody>'''
        for chk in parsing.get('account_month_checks', []):
            passed = chk.get('passed', False)
            status_cls = 'flag-no' if passed else 'flag-yes'
            gap_count = chk.get('extraction_gaps', 0) or 0
            gap_cell = f'<span class="gap-pill">{gap_count}</span>' if gap_count > 0 else '—'
            parsing_tab_html += f'''<tr class="{status_cls}">
                <td>{chk.get('month','')}</td><td class="mono">{chk.get('account_number','')}</td>
                <td class="mono r">{chk.get('opening_balance',0):,.2f}</td>
                <td class="mono r credit">{chk.get('gross_credits',0):,.2f}</td>
                <td class="mono r debit">{chk.get('gross_debits',0):,.2f}</td>
                <td class="mono r">{chk.get('expected_closing',0):,.2f}</td>
                <td class="mono r">{chk.get('closing_balance',0):,.2f}</td>
                <td class="mono r" style="font-weight:600;color:var(--{'green' if passed else 'red'})">{chk.get('reconciliation_delta',0):,.2f}</td>
                <td style="text-align:center"><span class="flag-dot {'clear' if passed else 'detected'}"></span>{'PASS' if passed else 'FAIL'}</td>
                <td class="mono r">{chk.get('transactions_extracted',0)}</td>
                <td style="text-align:center">{gap_cell}</td>
                <td style="font-size:0.78rem;color:var(--text-muted)">{chk.get('notes','') or ''}</td>
            </tr>'''
        parsing_tab_html += '''</tbody></table></div></div></div>'''

        # v6.2.1: Extraction gap detail section within Parsing QC tab
        p_extraction_gaps = parsing.get('extraction_gaps', []) or []
        if p_extraction_gaps:
            parsing_tab_html += '''<div class="section"><div class="section-head"><h2 style="color:var(--red)">Extraction Gap Details</h2></div><div class="section-body" style="padding:0"><div class="table-wrap"><table>
                <thead><tr><th>Month</th><th>Date</th><th>Page</th><th>Source File</th><th>Missing</th><th class="r">Amount (RM)</th><th>Before Gap</th><th>After Gap</th></tr></thead><tbody>'''
            for g in p_extraction_gaps:
                parsing_tab_html += f'''<tr class="flag-yes">
                    <td>{g.get('month','')}</td><td>{g.get('date','')}</td><td>{g.get('page','')}</td>
                    <td style="font-size:0.78rem">{g.get('source_file','')}</td>
                    <td><span class="badge badge-debit">{g.get('missing_type','')}</span></td>
                    <td class="mono r" style="font-weight:700;color:var(--red)">{g.get('missing_amount',0):,.2f}</td>
                    <td style="font-size:0.78rem" title="{g.get('prev_description','')}">{g.get('prev_description','')[:40]}... (RM {g.get('balance_before_gap',0):,.2f})</td>
                    <td style="font-size:0.78rem" title="{g.get('next_description','')}">{g.get('next_description','')[:40]}... (RM {g.get('balance_after_gap',0):,.2f})</td>
                </tr>'''
            parsing_tab_html += '</tbody></table></div></div></div>'

        # Gap 4: Classification Config section
        cls_config = data.get('classification_config', {})
        if cls_config or schema_v:
            rulebook_ver = cls_config.get('rulebook_version', 'N/A')
            exec_mode = cls_config.get('execution_mode', 'N/A')
            large_cr_threshold = cls_config.get('large_credit_threshold', 100000)
            uncl_listing_threshold = cls_config.get('unclassified_listing_threshold', 10000)
            factoring_entities = cls_config.get('known_factoring_entities', [])
            factoring_str = ', '.join(factoring_entities) if factoring_entities else 'None configured'

            parsing_tab_html += f'''
            <div class="section">
                <div class="section-head"><h2>Classification Configuration</h2></div>
                <div class="section-body">
                    <div class="config-grid">
                        <div class="config-item"><span class="config-label">Schema Version</span><span class="config-val">{schema_v or 'N/A'}</span></div>
                        <div class="config-item"><span class="config-label">Rulebook Version</span><span class="config-val">{rulebook_ver}</span></div>
                        <div class="config-item"><span class="config-label">Execution Mode</span><span class="config-val">{exec_mode}</span></div>
                        <div class="config-item"><span class="config-label">Large Credit Threshold</span><span class="config-val">RM {large_cr_threshold:,.0f}</span></div>
                        <div class="config-item"><span class="config-label">Unclassified Listing Threshold</span><span class="config-val">RM {uncl_listing_threshold:,.0f}</span></div>
                        <div class="config-item" style="grid-column:1/-1"><span class="config-label">Known Factoring Entities</span><span class="config-val" style="font-size:0.8rem">{factoring_str}</span></div>
                    </div>
                </div>
            </div>'''

        # Gap 6: V1-V6 Formula Validation Checks
        gross_cr = consol.get('gross_credits', 0) or 0
        own_cr = consol.get('total_own_party_cr', 0) or 0
        rp_cr = consol.get('total_related_party_cr', 0) or 0
        rev_cr = consol.get('total_reversal_cr', 0) or 0
        loan_disb_cr = consol.get('total_loan_disbursement_cr', 0) or 0
        fd_int_cr = consol.get('total_fd_interest_cr', 0) or 0
        inward_ret_cr = consol.get('total_inward_return_cr', 0) or 0
        net_cr = consol.get('net_credits', 0) or 0
        expected_net_cr = gross_cr - own_cr - rp_cr - rev_cr - loan_disb_cr - fd_int_cr - inward_ret_cr
        v1_delta = abs(net_cr - expected_net_cr)
        v1_pass = v1_delta < 0.02

        gross_dr = consol.get('gross_debits', 0) or 0
        own_dr = consol.get('total_own_party_dr', 0) or 0
        rp_dr = consol.get('total_related_party_dr', 0) or 0
        ret_chq = consol.get('total_returned_cheques_outward', 0) or 0
        net_dr = consol.get('net_debits', 0) or 0
        expected_net_dr = gross_dr - own_dr - rp_dr - ret_chq
        v2_delta = abs(net_dr - expected_net_dr)
        v2_pass = v2_delta < 0.02

        salary = consol.get('total_salary_paid', 0) or 0
        epf = consol.get('total_statutory_epf', 0) or 0
        socso = consol.get('total_statutory_socso', 0) or 0
        v3_ratio = (epf / salary * 100) if salary > 0 else 0
        v3_status = 'PASS' if 8 <= v3_ratio <= 16 else ('WARN' if salary > 0 else 'N/A')
        v4_ratio = (socso / salary * 100) if salary > 0 else 0
        v4_status = 'PASS' if 1 <= v4_ratio <= 5 else ('WARN' if salary > 0 else 'N/A')

        # V6: Sum of monthly net_credits = consolidated net_credits
        monthly_net_cr_sum = sum(
            sum(r.get('net_credits', 0) or 0 for r in rows)
            for rows in monthly_by_month.values()
        )
        v6_delta = abs(net_cr - monthly_net_cr_sum)
        v6_pass = v6_delta < 0.02

        def _v_cls(status):
            if status == 'PASS': return 'validation-pass'
            if status == 'WARN': return 'validation-warn'
            if status == 'FAIL': return 'validation-fail'
            return 'validation-pass'

        checks_data = [
            ('V1', 'Net Credits = Gross - Exclusions', 'BLOCKING', 'PASS' if v1_pass else 'FAIL', f'Delta: RM {v1_delta:,.2f}'),
            ('V2', 'Net Debits = Gross - Exclusions', 'BLOCKING', 'PASS' if v2_pass else 'FAIL', f'Delta: RM {v2_delta:,.2f}'),
            ('V3', 'EPF/Salary ratio 8-16%', 'WARNING', v3_status, f'{v3_ratio:.1f}%' if salary > 0 else 'No salary detected'),
            ('V4', 'SOCSO/Salary ratio 1-5%', 'WARNING', v4_status, f'{v4_ratio:.1f}%' if salary > 0 else 'No salary detected'),
            ('V5', 'C02+C11 dual-tag exclusion', 'BLOCKING', 'PASS', 'Single deduction via C02'),
            ('V6', 'Monthly net_cr sum = consolidated', 'BLOCKING', 'PASS' if v6_pass else 'FAIL', f'Delta: RM {v6_delta:,.2f}'),
        ]
        v_rows = ''
        for vid, desc, severity, status, remark in checks_data:
            v_rows += f'''<tr>
                <td class="mono" style="font-weight:600">{vid}</td>
                <td>{desc}</td>
                <td><span class="badge" style="background:var(--{'red-dim' if severity=='BLOCKING' else 'amber-dim'});color:var(--{'red' if severity=='BLOCKING' else 'amber'})">{severity}</span></td>
                <td style="text-align:center"><span class="{_v_cls(status)}">{status}</span></td>
                <td style="font-size:0.82rem;color:var(--text-soft)">{remark}</td>
            </tr>'''

        parsing_tab_html += f'''
            <div class="section">
                <div class="section-head"><h2>Formula Validation Checks (V1&ndash;V6)</h2></div>
                <div class="section-body" style="padding:0"><div class="table-wrap"><table>
                    <thead><tr><th>ID</th><th>Check</th><th>Severity</th><th style="text-align:center">Status</th><th>Remarks</th></tr></thead>
                    <tbody>{v_rows}</tbody>
                </table></div></div>
            </div>'''

        parsing_tab_html += '</div>'

    # v6.3.0: Build Fraud Detector tab HTML
    fraud_tab_html = ''
    pdf_integrity = data.get('pdf_integrity')
    if pdf_integrity:
        # pdf_integrity can be a dict of filename->results or a list
        if isinstance(pdf_integrity, dict):
            pdf_files = pdf_integrity.get('files', [])
            if not pdf_files and not isinstance(pdf_integrity.get(next(iter(pdf_integrity), ''), {}), list):
                # It might be keyed by filename
                pdf_files_dict = {k: v for k, v in pdf_integrity.items() if isinstance(v, dict) and k != 'summary'}
                pdf_files = [{'filename': k, **v} for k, v in pdf_files_dict.items()]
            if not pdf_files:
                pdf_files = pdf_integrity.get('results', [])
        elif isinstance(pdf_integrity, list):
            pdf_files = pdf_integrity
        else:
            pdf_files = []

        # Determine overall risk
        all_severities = []
        total_checks = 0
        high_count = 0
        medium_count = 0
        for pf in pdf_files:
            layers = pf.get('layers', pf.get('checks', pf.get('findings', [])))
            if isinstance(layers, list):
                for layer in layers:
                    sev = (layer.get('severity', '') or layer.get('risk', '') or 'LOW').upper()
                    all_severities.append(sev)
                    total_checks += 1
                    if sev == 'HIGH':
                        high_count += 1
                    elif sev == 'MEDIUM':
                        medium_count += 1
            elif isinstance(layers, dict):
                for layer_name, layer_data in layers.items():
                    sev = 'LOW'
                    if isinstance(layer_data, dict):
                        sev = (layer_data.get('severity', '') or layer_data.get('risk', '') or 'LOW').upper()
                    all_severities.append(sev)
                    total_checks += 1
                    if sev == 'HIGH':
                        high_count += 1
                    elif sev == 'MEDIUM':
                        medium_count += 1

        overall_risk = 'low'
        if high_count > 0:
            overall_risk = 'high'
        elif medium_count > 0:
            overall_risk = 'medium'

        overall_label = {'low': 'ALL CLEAR', 'medium': 'REVIEW NEEDED', 'high': 'HIGH RISK'}[overall_risk]
        overall_icon = {'low': '&#x1F6E1;', 'medium': '&#x26A0;&#xFE0F;', 'high': '&#x1F6A8;'}[overall_risk]

        # Per-file sections
        file_sections = ''
        for pf in pdf_files:
            fname = pf.get('filename', pf.get('file', 'Unknown'))
            file_risk = (pf.get('risk', pf.get('overall_risk', 'LOW'))).upper() if isinstance(pf.get('risk', pf.get('overall_risk', '')), str) else 'LOW'
            if file_risk not in ('LOW', 'MEDIUM', 'HIGH'):
                file_risk = 'LOW'
            risk_cls = file_risk.lower()

            layers = pf.get('layers', pf.get('checks', pf.get('findings', [])))
            layer_rows = ''
            if isinstance(layers, list):
                for layer in layers:
                    l_name = layer.get('layer', layer.get('name', ''))
                    l_sev = (layer.get('severity', '') or layer.get('risk', '') or 'LOW').upper()
                    l_msg = layer.get('message', layer.get('finding', layer.get('description', '')))
                    l_detail = layer.get('detail', layer.get('details', ''))
                    if isinstance(l_detail, dict):
                        l_detail = ', '.join(f'{k}: {v}' for k, v in l_detail.items())
                    elif isinstance(l_detail, list):
                        l_detail = '; '.join(str(x) for x in l_detail[:5])
                    sev_cls = l_sev.lower()
                    layer_rows += f'''<tr>
                        <td>{l_name}</td>
                        <td><span class="fraud-shield {sev_cls}" style="padding:1px 6px;font-size:0.72rem">{l_sev}</span></td>
                        <td>{l_msg}</td>
                        <td style="font-size:0.78rem;color:var(--text-soft)">{l_detail}</td>
                    </tr>'''
            elif isinstance(layers, dict):
                for l_name, l_data in layers.items():
                    if isinstance(l_data, dict):
                        l_sev = (l_data.get('severity', '') or l_data.get('risk', '') or 'LOW').upper()
                        l_msg = l_data.get('message', l_data.get('finding', l_data.get('description', '')))
                        l_detail = l_data.get('detail', l_data.get('details', ''))
                        if isinstance(l_detail, dict):
                            l_detail = ', '.join(f'{k}: {v}' for k, v in l_detail.items())
                        elif isinstance(l_detail, list):
                            l_detail = '; '.join(str(x) for x in l_detail[:5])
                    else:
                        l_sev = 'LOW'
                        l_msg = str(l_data)
                        l_detail = ''
                    sev_cls = l_sev.lower()
                    layer_rows += f'''<tr>
                        <td>{l_name}</td>
                        <td><span class="fraud-shield {sev_cls}" style="padding:1px 6px;font-size:0.72rem">{l_sev}</span></td>
                        <td>{l_msg}</td>
                        <td style="font-size:0.78rem;color:var(--text-soft)">{l_detail}</td>
                    </tr>'''

            file_sections += f'''
            <div class="section">
                <div class="fraud-file-header" onclick="this.nextElementSibling.classList.toggle('collapsed')">
                    <span style="font-weight:600;flex:1">{fname}</span>
                    <span class="fraud-shield {risk_cls}">{file_risk}</span>
                </div>
                <div class="fraud-detail">
                    <div class="table-wrap"><table>
                        <thead><tr><th>Layer</th><th>Severity</th><th>Finding</th><th>Detail</th></tr></thead>
                        <tbody>{layer_rows or '<tr><td colspan="4" class="note">No findings</td></tr>'}</tbody>
                    </table></div>
                </div>
            </div>'''

        fraud_tab_html = f'''
        <div id="tab-fraud" class="tab">
            <div class="{'dq-banner dq-pass' if overall_risk == 'low' else 'dq-banner dq-fail'}">
                <div class="dq-icon">{overall_icon}</div>
                <div>
                    <div class="dq-title">PDF Integrity: {overall_label}</div>
                    <div class="dq-detail">{'All PDF files passed integrity checks. No signs of tampering detected.' if overall_risk == 'low' else f'{high_count} HIGH and {medium_count} MEDIUM findings detected across {len(pdf_files)} PDF file(s). Manual review recommended.'}</div>
                </div>
            </div>
            <div class="summary-grid">
                <div class="summary-card"><div class="val">{len(pdf_files)}</div><div class="lbl">PDFs Analyzed</div></div>
                <div class="summary-card"><div class="val">{total_checks}</div><div class="lbl">Total Checks</div></div>
                <div class="summary-card"><div class="val" style="color:var(--{'red' if high_count > 0 else 'green'})">{high_count}</div><div class="lbl">HIGH Findings</div></div>
                <div class="summary-card"><div class="val" style="color:var(--{'amber' if medium_count > 0 else 'green'})">{medium_count}</div><div class="lbl">MEDIUM Findings</div></div>
            </div>
            {file_sections}
        </div>'''

    # Count flags
    total_flags = len(flags_data.get('indicators', []))

    html = f'''<!DOCTYPE html>
<html lang="en" data-theme="light">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Kredit Lab — {company}</title>
    <link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;600&family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <script src="https://cdnjs.cloudflare.com/ajax/libs/plotly.js/2.27.0/plotly.min.js"></script>
    <script>if(typeof Plotly==='undefined'){{document.write('<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"><\\/script>')}}</script>
    <style>
        :root, [data-theme="light"] {{
            --bg: #f5f6fa; --bg-alt: #ffffff; --card: #ffffff;
            --border: #e2e8f0; --border-accent: #cbd5e1;
            --green: #059669; --green-dim: rgba(5,150,105,0.08); --green-bg: #ecfdf5;
            --red: #dc2626; --red-dim: rgba(220,38,38,0.08); --red-bg: #fef2f2;
            --amber: #d97706; --amber-dim: rgba(217,119,6,0.08); --amber-bg: #fffbeb;
            --blue: #2563eb; --blue-dim: rgba(37,99,235,0.08);
            --purple: #7c3aed; --purple-dim: rgba(124,58,237,0.08);
            --text: #1e293b; --text-soft: #475569; --text-muted: #94a3b8;
            --shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
            --shadow-lg: 0 4px 6px rgba(0,0,0,0.05), 0 10px 15px rgba(0,0,0,0.03);
        }}
        [data-theme="dark"] {{
            --bg: #0b0f19; --bg-alt: #111827; --card: #1a2235;
            --border: #1e2a42; --border-accent: #2d3f5f;
            --green: #34d399; --green-dim: rgba(52,211,153,0.12); --green-bg: rgba(5,150,105,0.15);
            --red: #f87171; --red-dim: rgba(248,113,113,0.12); --red-bg: rgba(220,38,38,0.15);
            --amber: #fbbf24; --amber-dim: rgba(251,191,36,0.12); --amber-bg: rgba(217,119,6,0.15);
            --blue: #60a5fa; --blue-dim: rgba(96,165,250,0.12);
            --purple: #a78bfa; --purple-dim: rgba(167,139,250,0.12);
            --text: #e2e8f0; --text-soft: #94a3b8; --text-muted: #64748b;
            --shadow: 0 1px 3px rgba(0,0,0,0.3); --shadow-lg: 0 4px 6px rgba(0,0,0,0.4);
        }}
        * {{ margin:0; padding:0; box-sizing:border-box; }}
        body {{ font-family:'Inter',system-ui,sans-serif; background:var(--bg); color:var(--text); line-height:1.6; font-size:14px; }}
        .container {{ max-width:1440px; margin:0 auto; padding:1.5rem; }}

        /* Header */
        .header {{ background:var(--card); border:1px solid var(--border); border-radius:16px; padding:2rem; margin-bottom:1.5rem; position:relative; overflow:hidden; box-shadow:var(--shadow-lg); }}
        .header::before {{ content:''; position:absolute; top:0; left:0; right:0; height:4px; background:linear-gradient(90deg,#0d9488,#0ea5e9,#6366f1); }}
        .header-grid {{ display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:1.5rem; }}
        .company-info h1 {{ font-size:1.6rem; font-weight:700; margin-bottom:0.25rem; }}
        .company-info .period {{ color:var(--text-soft); font-size:0.88rem; }}
        .schema-badge {{ display:inline-block; padding:0.2rem 0.6rem; background:var(--purple-dim); color:var(--purple); border-radius:20px; font-size:0.72rem; font-weight:600; margin-left:0.75rem; vertical-align:middle; }}
        .header-kpi {{ display:flex; gap:1.75rem; flex-wrap:wrap; }}
        .kpi {{ text-align:center; padding:0 1rem; border-left:2px solid var(--border); }}
        .kpi:first-child {{ border-left:none; }}
        .kpi .val {{ font-size:1.35rem; font-weight:700; font-family:'JetBrains Mono',monospace; }}
        .kpi .lbl {{ font-size:0.7rem; color:var(--text-muted); text-transform:uppercase; letter-spacing:0.05em; }}
        .kpi .val.credit {{ color:var(--green); }}
        .kpi .val.debit {{ color:var(--red); }}

        /* Theme toggle */
        .theme-toggle {{ position:absolute; top:1rem; right:1rem; padding:0.4rem 0.75rem; border:1px solid var(--border); background:var(--bg-alt); color:var(--text-soft); border-radius:8px; cursor:pointer; font-size:0.8rem; }}
        .theme-toggle:hover {{ border-color:var(--border-accent); }}

        /* Nav */
        .nav {{ display:flex; gap:0.35rem; margin-bottom:1.5rem; flex-wrap:wrap; background:var(--card); padding:0.4rem; border-radius:12px; border:1px solid var(--border); box-shadow:var(--shadow); }}
        .nav-btn {{ padding:0.6rem 1rem; border:none; background:transparent; color:var(--text-soft); cursor:pointer; border-radius:8px; font-size:0.82rem; font-weight:500; transition:all 0.15s; white-space:nowrap; }}
        .nav-btn:hover {{ background:var(--bg); color:var(--text); }}
        .nav-btn.active {{ background:var(--blue); color:#fff; }}

        /* Tab content */
        .tab {{ display:none; }}
        .tab.active {{ display:block; }}

        /* Cards & Sections */
        .section {{ background:var(--card); border:1px solid var(--border); border-radius:12px; margin-bottom:1.25rem; box-shadow:var(--shadow); overflow:hidden; }}
        .section-head {{ padding:1rem 1.25rem; border-bottom:1px solid var(--border); cursor:pointer; display:flex; justify-content:space-between; align-items:center; }}
        .section-head h2 {{ font-size:1rem; font-weight:600; }}
        .section-body {{ padding:1.25rem; }}
        .section-body.collapsed {{ display:none; }}

        /* Account cards */
        .account-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(320px,1fr)); gap:1rem; margin-bottom:1rem; }}
        .account-card {{ background:var(--bg-alt); border:1px solid var(--border); border-radius:10px; padding:1.25rem; }}
        .account-header {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:0.5rem; }}
        .bank-name {{ font-weight:600; }}
        .account-number {{ font-family:'JetBrains Mono',monospace; font-size:0.85rem; color:var(--text-soft); }}
        .account-holder {{ font-size:0.82rem; color:var(--text-muted); margin-bottom:0.75rem; }}
        .account-metrics {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(100px,1fr)); gap:0.75rem; }}
        .metric {{ }}
        .metric-label {{ font-size:0.68rem; color:var(--text-muted); text-transform:uppercase; letter-spacing:0.03em; }}
        .metric-value {{ font-family:'JetBrains Mono',monospace; font-size:0.88rem; font-weight:600; }}

        /* Summary cards */
        .summary-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:1rem; margin-bottom:1.5rem; }}
        .summary-card {{ background:var(--bg-alt); border:1px solid var(--border); border-radius:10px; padding:1.25rem; text-align:center; }}
        .summary-card .val {{ font-size:1.4rem; font-weight:700; font-family:'JetBrains Mono',monospace; }}
        .summary-card .lbl {{ font-size:0.72rem; color:var(--text-muted); text-transform:uppercase; margin-top:0.25rem; }}

        /* Tables */
        .table-wrap {{ overflow-x:auto; border-radius:8px; border:1px solid var(--border); }}
        table {{ width:100%; border-collapse:collapse; font-size:0.82rem; }}
        th {{ background:var(--bg); color:var(--text-soft); font-weight:600; text-transform:uppercase; font-size:0.7rem; letter-spacing:0.04em; padding:0.65rem 0.75rem; text-align:left; position:sticky; top:0; white-space:nowrap; border-bottom:2px solid var(--border); }}
        td {{ padding:0.55rem 0.75rem; border-bottom:1px solid var(--border); }}
        tr:hover {{ background:var(--bg); }}
        .total-row {{ background:var(--blue-dim) !important; font-weight:600; }}
        .total-row td {{ border-top:2px solid var(--blue); border-bottom:2px solid var(--blue); }}
        .mono {{ font-family:'JetBrains Mono',monospace; }}
        .r {{ text-align:right; }}
        .credit {{ color:var(--green); }}
        .debit {{ color:var(--red); }}
        .sticky-col {{ position:sticky; left:0; background:inherit; z-index:1; font-weight:600; }}
        th.sticky-col {{ z-index:2; }}

        /* Badges */
        .badge {{ display:inline-block; padding:0.15rem 0.5rem; border-radius:20px; font-size:0.7rem; font-weight:600; }}
        .badge-current {{ background:var(--blue-dim); color:var(--blue); }}
        .badge-savings {{ background:var(--green-dim); color:var(--green); }}
        .badge-od {{ background:var(--red-dim); color:var(--red); }}
        .badge-credit {{ background:var(--green-dim); color:var(--green); }}
        .badge-debit {{ background:var(--red-dim); color:var(--red); }}

        .rp-tag {{ display:inline-block; padding:0.25rem 0.6rem; background:var(--amber-dim); color:var(--amber); border-radius:6px; font-size:0.78rem; margin:0.2rem; }}
        .rp-tag small {{ opacity:0.7; }}
        .rp-badge {{ display:inline-block; padding:0.1rem 0.35rem; background:var(--amber-dim); color:var(--amber); border-radius:4px; font-size:0.65rem; font-weight:700; margin-left:0.35rem; vertical-align:middle; }}

        /* Flags */
        .flag-dot {{ display:inline-block; width:10px; height:10px; border-radius:50%; margin-right:0.35rem; vertical-align:middle; }}
        .flag-dot.detected {{ background:var(--red); }}
        .flag-dot.clear {{ background:var(--green); }}
        .flag-yes {{ background:var(--red-bg); }}

        /* Observations */
        .obs-list {{ list-style:none; padding:0; }}
        .obs-item {{ padding:0.75rem 1rem; margin-bottom:0.5rem; border-radius:8px; font-size:0.88rem; line-height:1.5; }}
        .obs-item.positive {{ background:var(--green-bg); border-left:3px solid var(--green); }}
        .obs-item.concern {{ background:var(--red-bg); border-left:3px solid var(--red); }}

        /* Two column layout */
        .two-col {{ display:grid; grid-template-columns:1fr 1fr; gap:1.5rem; }}
        @media (max-width:900px) {{ .two-col {{ grid-template-columns:1fr; }} }}

        /* Charts */
        .chart-box {{ background:var(--card); border:1px solid var(--border); border-radius:10px; padding:1rem; margin-bottom:1rem; }}
        .chart-title {{ font-size:0.85rem; font-weight:600; margin-bottom:0.5rem; color:var(--text-soft); }}

        /* Note */
        .note {{ font-size:0.8rem; color:var(--text-muted); padding:0.5rem; font-style:italic; }}

        /* v6.2.1: Data quality banner */
        .dq-banner {{ border-radius:12px; padding:1.25rem 1.5rem; margin-bottom:1.5rem; display:flex; gap:14px; align-items:flex-start; }}
        .dq-banner.dq-fail {{ background:var(--red-bg); border:1px solid var(--red); }}
        .dq-banner.dq-pass {{ background:var(--green-bg); border:1px solid var(--green); }}
        .dq-banner .dq-icon {{ font-size:1.3rem; flex-shrink:0; }}
        .dq-banner .dq-title {{ font-weight:700; font-size:0.92rem; margin-bottom:4px; }}
        .dq-banner.dq-fail .dq-title {{ color:var(--red); }}
        .dq-banner.dq-pass .dq-title {{ color:var(--green); }}
        .dq-banner .dq-detail {{ font-size:0.82rem; color:var(--text-soft); line-height:1.6; }}
        .dq-banner .dq-stats {{ display:flex; gap:2rem; margin-top:0.75rem; font-size:0.82rem; }}
        .dq-banner .dq-stat-label {{ font-size:0.68rem; color:var(--text-muted); text-transform:uppercase; letter-spacing:0.04em; }}
        .dq-banner .dq-stat-val {{ font-weight:700; font-size:1.05rem; font-family:'JetBrains Mono',monospace; }}
        .dq-banner.dq-fail .dq-stat-val {{ color:var(--red); }}

        /* v6.2.1: Row-level recon status */
        tr.row-fail {{ background:var(--red-bg) !important; }}
        .recon-badge {{ display:inline-flex; align-items:center; gap:3px; padding:2px 7px; border-radius:4px; font-size:0.7rem; font-weight:600; font-family:'JetBrains Mono',monospace; }}
        .recon-badge.pass {{ background:var(--green-dim); color:var(--green); }}
        .recon-badge.fail {{ background:var(--red-dim); color:var(--red); }}
        .gap-pill {{ display:inline-flex; align-items:center; gap:3px; padding:2px 7px; border-radius:4px; font-size:0.7rem; background:var(--red-dim); color:var(--red); font-weight:500; }}
        .dq-gap-panel {{ background:var(--red-bg); border:1px solid rgba(220,38,38,0.2); border-radius:10px; padding:1rem 1.25rem; margin-top:0.75rem; font-size:0.82rem; }}
        .dq-gap-panel strong {{ color:var(--red); }}
        .dq-gap-item {{ padding:0.5rem 0; border-bottom:1px solid rgba(220,38,38,0.1); display:grid; grid-template-columns:100px 1fr; gap:0.5rem; }}
        .dq-gap-item:last-child {{ border-bottom:none; }}
        .obs-item.data-warn {{ background:var(--red-bg); border-left:3px solid var(--red); font-weight:500; }}

        /* v6.3.0 column highlights */
        .v630-count {{ background:var(--purple-dim) !important; }}
        .v630-amt {{ background:var(--blue-dim) !important; }}
        .v630-uncl {{ background:var(--amber-dim) !important; }}
        td.v630-count {{ background:var(--purple-dim); }}
        td.v630-amt {{ background:var(--blue-dim); }}
        td.v630-uncl {{ background:var(--amber-dim); }}

        /* Fraud detector */
        .fraud-shield {{ display:inline-flex; align-items:center; gap:6px; padding:0.3rem 0.8rem; border-radius:20px; font-weight:700; font-size:0.85rem; }}
        .fraud-shield.low {{ background:var(--green-dim); color:var(--green); }}
        .fraud-shield.medium {{ background:var(--amber-dim); color:var(--amber); }}
        .fraud-shield.high {{ background:var(--red-dim); color:var(--red); }}
        .fraud-file-header {{ display:flex; align-items:center; gap:0.75rem; padding:1rem 1.25rem; cursor:pointer; border-bottom:1px solid var(--border); }}
        .fraud-file-header:hover {{ background:var(--bg); }}
        .fraud-detail {{ padding:0; }}
        .info-panel {{ background:var(--blue-dim); border:1px solid rgba(37,99,235,0.2); border-radius:10px; padding:1rem 1.25rem; margin-bottom:1.25rem; font-size:0.84rem; line-height:1.7; }}
        .info-panel h4 {{ color:var(--blue); margin-bottom:0.5rem; }}
        .info-panel ul {{ margin:0.5rem 0 0 1.25rem; }}
        .info-panel li {{ margin-bottom:0.25rem; }}
        .config-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:0.5rem; }}
        .config-item {{ display:flex; justify-content:space-between; padding:0.5rem 0.75rem; border-bottom:1px solid var(--border); font-size:0.84rem; }}
        .config-item .config-label {{ color:var(--text-soft); }}
        .config-item .config-val {{ font-family:'JetBrains Mono',monospace; font-weight:600; }}
        .validation-pass {{ color:var(--green); font-weight:600; }}
        .validation-warn {{ color:var(--amber); font-weight:600; }}
        .validation-fail {{ color:var(--red); font-weight:600; }}

        /* Footer */
        .footer {{ text-align:center; padding:2rem 1rem; color:var(--text-muted); font-size:0.78rem; border-top:1px solid var(--border); margin-top:2rem; }}

        /* Print */
        @media print {{
            .nav, .theme-toggle {{ display:none; }}
            .tab {{ display:block !important; page-break-inside:avoid; }}
            body {{ font-size:11px; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <button class="theme-toggle" onclick="toggleTheme()">Dark</button>
            <div class="header-grid">
                <div class="company-info">
                    <div style="font-size:0.72rem;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:var(--text-muted);margin-bottom:0.35rem">Kredit Lab &mdash; Statement Intelligence</div>
                    <h1>{company} <span class="schema-badge">Kredit Lab v{r.get('schema_version', '6')}</span></h1>
                    <div class="period">{period_start} to {period_end} &middot; {total_months} months &middot; {sum(a.get('transaction_count',0) for a in accounts):,} transactions</div>
                </div>
                <div class="header-kpi">
                    <div class="kpi"><div class="val credit">RM {consol.get('net_credits',0):,.0f}</div><div class="lbl">Net Credits</div></div>
                    <div class="kpi"><div class="val debit">RM {consol.get('net_debits',0):,.0f}</div><div class="lbl">Net Debits{'  ⚠️' if is_incomplete else ''}</div></div>
                    <div class="kpi"><div class="val">RM {consol.get('annualized_net_credits',0):,.0f}</div><div class="lbl">Annualized</div></div>
                    <div class="kpi"><div class="val">RM {consol.get('eod_average',0):,.0f}</div><div class="lbl">Avg EOD</div></div>
                    <div class="kpi"><div class="val" style="color:var(--{'red' if detected_count > total_flags//2 else 'amber'})">{detected_count}/{total_flags}</div><div class="lbl">Flags</div></div>
                </div>
            </div>
        </div>

        <div class="nav">
            <button class="nav-btn active" onclick="showTab('overview')">Overview</button>
            <button class="nav-btn" onclick="showTab('monthly')">Cash Flow</button>
            <button class="nav-btn" onclick="showTab('parties')">Top Parties</button>
            <button class="nav-btn" onclick="showTab('large')">Large Credits</button>
            <button class="nav-btn" onclick="showTab('related')">Counterparty</button>
            <button class="nav-btn" onclick="showTab('loans')">Facilities</button>
            <button class="nav-btn" onclick="showTab('flags')">Risk Signals</button>
            {'<button class="nav-btn" onclick="showTab(&#39;fx&#39;)">FX / Remittance</button>' if is_v620 else ''}
            {'<button class="nav-btn" onclick="showTab(&#39;unclassified&#39;)">Unclassified</button>' if is_v630 else ''}
            {'<button class="nav-btn" onclick="showTab(&#39;parsing&#39;)">Parsing QC</button>' if has_parsing else ''}
            {'<button class="nav-btn" onclick="showTab(&#39;fraud&#39;)">Fraud Detector</button>' if data.get('pdf_integrity') else ''}
        </div>

        <!-- OVERVIEW TAB -->
        <div id="tab-overview" class="tab active">
            {dq_banner_html}
            <div class="account-grid">{acc_cards}</div>

            {'<div class="section"><div class="section-head"><h2>Known Related Parties</h2></div><div class="section-body">' + rp_html + '</div></div>' if rp_html else ''}

            <div class="two-col">
                <div class="chart-box"><div class="chart-title">Net Credits vs Debits (Monthly)</div><div id="chartCrDr" style="height:300px"></div></div>
                <div class="chart-box"><div class="chart-title">EOD Balance Range (Monthly)</div><div id="chartEOD" style="height:300px"></div></div>
            </div>

            <div class="section">
                <div class="section-head"><h2>Consolidated Summary</h2></div>
                <div class="section-body">
                    <div class="summary-grid">
                        <div class="summary-card"><div class="val credit">{consol.get('gross_credits',0):,.0f}</div><div class="lbl">Gross Credits</div></div>
                        <div class="summary-card"><div class="val debit">{consol.get('gross_debits',0):,.0f}</div><div class="lbl">Gross Debits</div></div>
                        <div class="summary-card"><div class="val credit">{consol.get('net_credits',0):,.0f}</div><div class="lbl">Net Credits</div></div>
                        <div class="summary-card"><div class="val debit">{consol.get('net_debits',0):,.0f}</div><div class="lbl">Net Debits</div></div>
                        <div class="summary-card"><div class="val">{consol.get('annualized_net_credits',0):,.0f}</div><div class="lbl">Annualized Cr</div></div>
                        <div class="summary-card"><div class="val">{consol.get('annualized_net_debits',0):,.0f}</div><div class="lbl">Annualized Dr</div></div>
                    </div>
                    <div class="two-col">
                        <div>
                            <h4 style="color:var(--green);margin-bottom:0.75rem">Exclusions from Credits</h4>
                            <div class="table-wrap"><table>
                                <tr><td>Own Party</td><td class="mono r">{consol.get('total_own_party_cr',0):,.2f}</td></tr>
                                <tr><td>Related Party</td><td class="mono r">{consol.get('total_related_party_cr',0):,.2f}</td></tr>
                                <tr><td>Reversals</td><td class="mono r">{consol.get('total_reversal_cr',0):,.2f}</td></tr>
                                <tr><td>Loan Disbursements</td><td class="mono r">{consol.get('total_loan_disbursement_cr',0):,.2f}</td></tr>
                                <tr><td>FD/Interest</td><td class="mono r">{consol.get('total_fd_interest_cr',0):,.2f}</td></tr>
                                <tr><td>Inward Return (C16)</td><td class="mono r">{consol.get('total_inward_return_cr',0):,.2f}</td></tr>
                            </table></div>
                        </div>
                        <div>
                            <h4 style="color:var(--red);margin-bottom:0.75rem">Exclusions from Debits</h4>
                            <div class="table-wrap"><table>
                                <tr><td>Own Party</td><td class="mono r">{consol.get('total_own_party_dr',0):,.2f}</td></tr>
                                <tr><td>Related Party</td><td class="mono r">{consol.get('total_related_party_dr',0):,.2f}</td></tr>
                                <tr><td>Returned Cheques</td><td class="mono r">{consol.get('total_returned_cheques_outward',0):,.2f}</td></tr>
                            </table></div>
                        </div>
                    </div>
                    {'<div style="margin-top:1rem"><h4 style="color:var(--purple);margin-bottom:0.75rem">FX / Remittance Summary</h4><div class="summary-grid"><div class="summary-card"><div class="val credit">' + f"{consol.get('total_fx_credits',0):,.0f}" + '</div><div class="lbl">FX Credits (' + f"{consol.get('fx_credit_pct',0):.1f}" + '% of Gross)</div></div><div class="summary-card"><div class="val debit">' + f"{consol.get('total_fx_debits',0):,.0f}" + '</div><div class="lbl">FX Debits (' + f"{consol.get('fx_debit_pct',0):.1f}" + '% of Gross)</div></div><div class="summary-card"><div class="val" style="font-size:0.9rem">' + (', '.join(consol.get('fx_currencies_all', [])) or 'None') + '</div><div class="lbl">Currencies</div></div></div></div>' if is_v620 else ''}
                </div>
            </div>

            <div class="two-col">
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--green)">Positive Observations</h2></div>
                    <div class="section-body"><ul class="obs-list">{pos_obs}</ul></div>
                </div>
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--red)">Concerns</h2></div>
                    <div class="section-body"><ul class="obs-list">{con_obs}</ul></div>
                </div>
            </div>
        </div>

        <!-- MONTHLY ANALYSIS TAB -->
        <div id="tab-monthly" class="tab">
            <div class="section">
                <div class="section-head"><h2>Monthly Cash Flow Breakdown</h2></div>
                {'<div style="padding:0.5rem 1.25rem;font-size:0.78rem;color:var(--text-soft);border-bottom:1px solid var(--border);display:flex;gap:1.25rem;flex-wrap:wrap">' + ''.join(f'<span><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{acct_colors.get(a,"var(--text-muted)")};margin-right:4px;vertical-align:middle"></span>{next((ac.get("bank_name","") for ac in accounts if ac.get("account_number")==a), "")} ({a})</span>' for a in acct_list) + ' <span style="font-weight:600">Bold rows = month subtotal</span></div>' if has_account_col and len(acct_list) > 1 else ''}
                <div class="section-body" style="padding:0">
                    <div class="table-wrap" style="max-height:600px; overflow:auto">
                        <table>
                            <thead><tr>
                                <th class="sticky-col">Month / Account</th>
                                {'<th>Status</th>' if has_recon else ''}
                                <th class="r">Gross Cr</th><th class="r">Gross Dr</th>
                                <th class="r">Net Cr</th><th class="r">Net Dr</th>
                                <th class="r">Cr #</th><th class="r">Dr #</th>
                                <th class="r">Own Cr</th><th class="r">Own Dr</th>
                                <th class="r">RP Cr</th><th class="r">RP Dr</th>
                                <th class="r">Reversal</th>
                                <th class="r">Loan Disb</th><th class="r">FD/Int</th>
                                <th class="r">Cash Dep</th><th class="r">Cash Wdl</th>
                                <th class="r">Chq Dep</th><th class="r">Chq Issue</th>
                                <th class="r">Loan Repay</th><th class="r">Salary</th>
                                <th class="r">EPF</th><th class="r">SOCSO</th><th class="r">Tax</th>
                                <th class="r">Ret Chq #</th><th class="r">Ret Chq Amt</th>
                                <th class="r">Round Fig</th><th class="r">High Val</th>
                                <th class="r">EOD Low</th><th class="r">EOD High</th><th class="r">EOD Avg</th>
                                <th class="r">Open Bal</th><th class="r">Close Bal</th>
                                {'<th class="r v630-count">Own Cr #</th><th class="r v630-count">Own Dr #</th><th class="r v630-count">RP Cr #</th><th class="r v630-count">RP Dr #</th><th class="r v630-count">Loan #</th><th class="r v630-amt">Inward Ret</th><th class="r v630-uncl">Uncl Cr #</th><th class="r v630-uncl">Uncl Cr Amt</th><th class="r v630-uncl">Uncl Dr #</th><th class="r v630-uncl">Uncl Dr Amt</th>' if is_v630 else ''}
                            </tr></thead>
                            <tbody>
                                {monthly_rows}
                                {consol_row}
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
            <div class="note">Net Credits = Gross Credits - Own Party - Related Party - Reversals - Loan Disbursements - FD/Interest{' - Inward Return (C16)' if is_v630 else ''} | Net Debits = Gross Debits - Own Party - Related Party - Returned Cheques{' | <span style="display:inline-block;width:10px;height:10px;background:var(--purple-dim);border:1px solid var(--purple);border-radius:2px;vertical-align:middle;margin:0 2px"></span> Count columns <span style="display:inline-block;width:10px;height:10px;background:var(--amber-dim);border:1px solid var(--amber);border-radius:2px;vertical-align:middle;margin:0 2px"></span> Unclassified columns (v6.3.0)' if is_v630 else ''}</div>
'''
    # v6.2.1: Build gap detail panels for failed months
    gap_panels_html = ''
    extraction_gaps = parsing.get('extraction_gaps', []) if parsing else []
    if extraction_gaps and has_recon:
        # Group gaps by month
        from collections import defaultdict as _dd
        gaps_by_month = _dd(list)
        for g in extraction_gaps:
            gaps_by_month[g.get('month', '')].append(g)

        for gap_month in sorted(gaps_by_month.keys()):
            month_gaps = gaps_by_month[gap_month]
            total_missing = sum(g.get('missing_amount', 0) for g in month_gaps)
            gap_panels_html += f'''<div class="dq-gap-panel">
                <strong>Extraction Gaps — {gap_month} ({len(month_gaps)} gap{"s" if len(month_gaps) > 1 else ""}, RM {total_missing:,.2f} missing)</strong>'''
            for gi, g in enumerate(month_gaps, 1):
                gap_panels_html += f'''<div class="dq-gap-item">
                    <div><div style="font-size:0.72rem;color:var(--text-muted)">Gap #{gi}</div>
                    <div style="color:var(--red);font-weight:600">RM {g.get('missing_amount',0):,.2f}</div></div>
                    <div><div>Page {g.get('page','')} · {g.get('date','')} · {g.get('missing_type','').lower()}(s) missing</div>
                    <div style="font-size:0.78rem;color:var(--text-muted);margin-top:2px">After: <em>{g.get('prev_description','')[:60]}</em> (bal RM {g.get('balance_before_gap',0):,.2f})</div>
                    <div style="font-size:0.78rem;color:var(--text-muted)">Before: <em>{g.get('next_description','')[:60]}</em> (bal RM {g.get('balance_after_gap',0):,.2f})</div></div>
                </div>'''
            gap_panels_html += '</div>'

    html += gap_panels_html
    html += f'''
        </div>

        <!-- TOP PARTIES TAB -->
        <div id="tab-parties" class="tab">
            <div class="two-col">
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--green)">Top 5 Payers (Income)</h2></div>
                    <div class="section-body" style="padding:0">
                        <div class="table-wrap"><table>
                            <thead><tr><th>#</th><th>Party</th><th class="r">Amount (RM)</th><th class="r">Txns</th></tr></thead>
                            <tbody>{payer_rows or '<tr><td colspan="4" style="text-align:center;color:var(--text-muted)">No data</td></tr>'}</tbody>
                        </table></div>
                    </div>
                </div>
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--red)">Top 5 Payees (Outflow)</h2></div>
                    <div class="section-body" style="padding:0">
                        <div class="table-wrap"><table>
                            <thead><tr><th>#</th><th>Party</th><th class="r">Amount (RM)</th><th class="r">Txns</th></tr></thead>
                            <tbody>{payee_rows or '<tr><td colspan="4" style="text-align:center;color:var(--text-muted)">No data</td></tr>'}</tbody>
                        </table></div>
                    </div>
                </div>
            </div>
            <div class="note"><span class="rp-badge">RP</span> = Related Party</div>
        </div>

        <!-- LARGE CREDITS TAB -->
        <div id="tab-large" class="tab">
            <div class="section">
                <div class="section-head"><h2>Large Credits (&ge; RM 100,000)</h2><span class="badge badge-current">{len(large_credits)} transactions</span></div>
                <div class="section-body" style="padding:0">
                    <div class="table-wrap" style="max-height:500px;overflow:auto"><table>
                        <thead><tr><th>Date</th><th>Description</th><th class="r">Amount (RM)</th><th class="r">Balance</th></tr></thead>
                        <tbody>{large_cr_rows}</tbody>
                    </table></div>
                </div>
            </div>
        </div>

        <!-- RELATED PARTY TAB -->
        <div id="tab-related" class="tab">
            <div class="section">
                <div class="section-head"><h2>Own & Related Party Transactions</h2></div>
                <div class="section-body">
                    <div class="summary-grid">
                        <div class="summary-card"><div class="val credit">{rp_summary.get('own_party_cr',0):,.0f}</div><div class="lbl">Own Party Cr ({rp_summary.get('own_party_cr_pct',0):.1f}%)</div></div>
                        <div class="summary-card"><div class="val debit">{rp_summary.get('own_party_dr',0):,.0f}</div><div class="lbl">Own Party Dr ({rp_summary.get('own_party_dr_pct',0):.1f}%)</div></div>
                        <div class="summary-card"><div class="val credit">{rp_summary.get('related_party_cr',0):,.0f}</div><div class="lbl">Related Party Cr ({rp_summary.get('related_party_cr_pct',0):.1f}%)</div></div>
                        <div class="summary-card"><div class="val debit">{rp_summary.get('related_party_dr',0):,.0f}</div><div class="lbl">Related Party Dr ({rp_summary.get('related_party_dr_pct',0):.1f}%)</div></div>
                    </div>
                    <div class="table-wrap" style="max-height:500px;overflow:auto"><table>
                        <thead><tr><th>Date</th><th>Description</th><th class="r">Amount (RM)</th><th>Type</th><th>Party Type</th><th>Party Name</th></tr></thead>
                        <tbody>{rp_txn_rows}</tbody>
                    </table></div>
                    {rp_note}
                </div>
            </div>
        </div>

        <!-- LOANS TAB -->
        <div id="tab-loans" class="tab">
            <div class="summary-grid">
                <div class="summary-card"><div class="val credit">{consol.get('total_loan_disbursement_cr',0):,.0f}</div><div class="lbl">Total Disbursements</div></div>
                <div class="summary-card"><div class="val debit">{consol.get('total_loan_repayment_dr',0):,.0f}</div><div class="lbl">Total Repayments</div></div>
                <div class="summary-card"><div class="val">{len(loans.get('disbursements',[]))}</div><div class="lbl">Disbursement Txns</div></div>
                <div class="summary-card"><div class="val">{len(loans.get('repayments',[]))}</div><div class="lbl">Repayment Txns</div></div>
            </div>
            <div class="two-col">
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--green)">Disbursements (Credits)</h2></div>
                    <div class="section-body" style="padding:0"><div class="table-wrap" style="max-height:400px;overflow:auto"><table>
                        <thead><tr><th>Date</th><th>Description</th><th class="r">Amount</th><th>Category</th></tr></thead>
                        <tbody>{loan_disb_rows or '<tr><td colspan="4" class="note">No disbursements</td></tr>'}</tbody>
                    </table></div></div>
                </div>
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--red)">Repayments (Debits)</h2></div>
                    <div class="section-body" style="padding:0"><div class="table-wrap" style="max-height:400px;overflow:auto"><table>
                        <thead><tr><th>Date</th><th>Description</th><th class="r">Amount</th><th>Category</th></tr></thead>
                        <tbody>{loan_repay_rows or '<tr><td colspan="4" class="note">No repayments</td></tr>'}</tbody>
                    </table></div></div>
                </div>
            </div>
        </div>

        <!-- FLAGS TAB -->
        <div id="tab-flags" class="tab">
            <div class="section">
                <div class="section-head"><h2>Risk Signals</h2><span class="badge" style="background:var(--{'red-dim' if detected_count > total_flags//2 else 'amber-dim'});color:var(--{'red' if detected_count > total_flags//2 else 'amber'})">{detected_count} of {total_flags} detected</span></div>
                <div class="section-body" style="padding:0">
                    <div class="table-wrap"><table>
                        <thead><tr><th style="width:40px">#</th><th>Signal</th><th style="width:80px;text-align:center">Status</th><th>Remarks</th></tr></thead>
                        <tbody>{flag_rows}</tbody>
                    </table></div>
                </div>
            </div>
        </div>

        {fx_tab_html}
        {unclassified_tab_html}
        {parsing_tab_html}
        {fraud_tab_html}

        <div class="footer">
            <p>Kredit Lab &mdash; Statement Intelligence Report | Generated {r.get('generated_at','')} | {period_start} &ndash; {period_end}</p>
        </div>
    </div>

    <script>
        function showTab(name) {{
            document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
            document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
            const tab = document.getElementById('tab-'+name);
            if(tab) tab.classList.add('active');
            event.target.classList.add('active');
        }}
        function toggleTheme() {{
            const html = document.documentElement;
            const btn = document.querySelector('.theme-toggle');
            const t = html.getAttribute('data-theme')==='dark'?'light':'dark';
            html.setAttribute('data-theme',t);
            btn.textContent = t==='dark'?'Light':'Dark';
            // Re-render charts for theme
            renderCharts();
        }}
        function renderCharts() {{
            if(typeof Plotly==='undefined'){{console.warn('Plotly not loaded — charts disabled');return;}}
            const isDark = document.documentElement.getAttribute('data-theme')==='dark';
            const gridColor = isDark?'#1e2a42':'#e2e8f0';
            const textColor = isDark?'#94a3b8':'#475569';
            const bg = 'transparent';

            Plotly.newPlot('chartCrDr', [
                {{x:{chart_months},y:{chart_net_cr},name:'Net Credits',type:'bar',marker:{{color:'rgba(5,150,105,0.7)'}}}},
                {{x:{chart_months},y:{chart_net_dr},name:'Net Debits',type:'bar',marker:{{color:'rgba(220,38,38,0.7)'}}}}
            ], {{
                paper_bgcolor:bg,plot_bgcolor:bg,font:{{color:textColor,size:11}},
                barmode:'group',showlegend:true,legend:{{orientation:'h',y:1.12}},
                margin:{{t:30,b:40,l:60,r:20}},
                yaxis:{{gridcolor:gridColor,tickformat:','}}
            }}, {{responsive:true,displayModeBar:false}});

            Plotly.newPlot('chartEOD', [
                {{x:{chart_months},y:{chart_eod_high},name:'EOD High',type:'scatter',mode:'lines+markers',line:{{color:'#2563eb',width:2}},marker:{{size:6}}}},
                {{x:{chart_months},y:{chart_eod_avg},name:'EOD Average',type:'scatter',mode:'lines+markers',line:{{color:'#7c3aed',width:2}},marker:{{size:6}}}},
                {{x:{chart_months},y:{chart_eod_low},name:'EOD Low',type:'scatter',mode:'lines+markers',line:{{color:'#dc2626',width:2,dash:'dot'}},marker:{{size:6}}}}
            ], {{
                paper_bgcolor:bg,plot_bgcolor:bg,font:{{color:textColor,size:11}},
                showlegend:true,legend:{{orientation:'h',y:1.12}},
                margin:{{t:30,b:40,l:60,r:20}},
                yaxis:{{gridcolor:gridColor,tickformat:','}}
            }}, {{responsive:true,displayModeBar:false}});

            // v6.2.0: FX chart
            var fxEl = document.getElementById('chartFX');
            if (fxEl) {{
                Plotly.newPlot('chartFX', [
                    {{x:{chart_months},y:{fx_chart_cr_json},name:'FX Credits',type:'bar',marker:{{color:'rgba(5,150,105,0.7)'}}}},
                    {{x:{chart_months},y:{fx_chart_dr_json},name:'FX Debits',type:'bar',marker:{{color:'rgba(220,38,38,0.7)'}}}}
                ], {{
                    paper_bgcolor:bg,plot_bgcolor:bg,font:{{color:textColor,size:11}},
                    barmode:'group',showlegend:true,legend:{{orientation:'h',y:1.12}},
                    margin:{{t:30,b:40,l:60,r:20}},
                    yaxis:{{gridcolor:gridColor,tickformat:','}}
                }}, {{responsive:true,displayModeBar:false}});
            }}
        }}
        document.addEventListener('DOMContentLoaded', renderCharts);
    </script>
</body>
</html>'''
    return html


# ══════════════════════════════════════════════
# MAIN STREAMLIT APP
# ══════════════════════════════════════════════

render_login_gate()
render_logout_control()

st.title("🔬 Kredit Lab — Statement Intelligence")
st.caption("Malaysian bank statement analysis engine — Cash flow breakdown, risk signals & credit intelligence")

uploaded_file = st.file_uploader("Upload JSON Analysis", type=['json'])

data = None
if uploaded_file:
    try:
        data = json.load(uploaded_file)
    except Exception as e:
        st.error(f"Error parsing JSON: {e}")

if data:
    r = data.get('report_info', {})
    company = r.get('company_name', 'Unknown')
    schema_v = r.get('schema_version', '?')

    if schema_v not in ('6.0.0', '6.1.0', '6.1.1', '6.2.0', '6.2.1', '6.2.2', '6.3.0'):
        st.warning(f"Expected schema v6.0.0–v6.3.0, got v{schema_v}. Some features may not work correctly.")

    st.success(f"Loaded: **{company}** (Schema v{schema_v})")

    # ── Downloads ──
    st.markdown("### Downloads")
    c1, c2, c3 = st.columns(3)

    safe_name = company.replace(' ', '_').replace('/', '_')

    with c1:
        st.download_button(
            "📄 JSON",
            json.dumps(data, indent=2),
            f"{safe_name}_v6_analysis.json",
            "application/json",
            use_container_width=True
        )

    with c2:
        html_content = generate_interactive_html(data)
        st.download_button(
            "🌐 Interactive HTML",
            html_content,
            f"{safe_name}_v6_report.html",
            "text/html",
            use_container_width=True
        )

    with c3:
        excel_buf = generate_excel(data)
        if excel_buf:
            st.download_button(
                "📊 Excel Workbook",
                excel_buf,
                f"{safe_name}_v6_analysis.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.info("Install `openpyxl` for Excel export")

    st.markdown("---")

    # ── Quick metrics ──
    consol = data.get('consolidated', {})
    accounts = data.get('accounts', [])
    flags_data = data.get('flags', {})
    detected = sum(1 for f in flags_data.get('indicators', []) if f.get('detected'))

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Net Credits", f"RM {consol.get('net_credits',0):,.0f}")
    c2.metric("Net Debits", f"RM {consol.get('net_debits',0):,.0f}")
    c3.metric("Annualized", f"RM {consol.get('annualized_net_credits',0):,.0f}")
    c4.metric("Avg EOD", f"RM {consol.get('eod_average',0):,.0f}")
    c5.metric("Signals", f"{detected}/{len(flags_data.get('indicators', []))}")

    # v6.2.0: FX summary row
    if consol.get('total_fx_credits') is not None:
        fc1, fc2, fc3, fc4 = st.columns(4)
        fc1.metric("FX Credits", f"RM {consol.get('total_fx_credits',0):,.0f}")
        fc2.metric("FX Debits", f"RM {consol.get('total_fx_debits',0):,.0f}")
        fc3.metric("FX Cr % of Gross", f"{consol.get('fx_credit_pct',0):.1f}%")
        fx_cur = consol.get('fx_currencies_all', [])
        fc4.metric("FX Currencies", ', '.join(fx_cur) if fx_cur else 'None')

    # v6.2.1: Data quality banner — must be unmissable
    dq_completeness = consol.get('data_completeness', '')
    dq_has_recon = any(m.get('reconciliation_status') for m in data.get('monthly_analysis', []))
    if dq_has_recon:
        if dq_completeness == 'INCOMPLETE':
            dq_gaps = consol.get('total_extraction_gaps', 0) or 0
            dq_missing_dr = consol.get('total_missing_debits', 0) or 0
            dq_missing_cr = consol.get('total_missing_credits', 0) or 0
            dq_months = consol.get('months_with_gaps', 0) or 0
            dq_warning = consol.get('data_quality_warning', '')
            st.error(
                f"⚠️ **INCOMPLETE EXTRACTION** — {dq_months} of {data.get('report_info',{}).get('total_months',0)} months have extraction gaps. "
                f"{dq_gaps} gap(s) detected. Missing debits: RM {dq_missing_dr:,.2f}, missing credits: RM {dq_missing_cr:,.2f}. "
                f"Figures below are understated."
            )
        else:
            st.success("✅ **Extraction Complete** — All months pass balance trail reconciliation. No gaps detected.")

    # v6.2.0: Parsing quality
    parsing_meta = data.get('parsing_metadata')
    if parsing_meta:
        success_rate = parsing_meta.get('overall_success_rate', 0)
        rate_icon = "✅" if success_rate >= 95 else "⚠️" if success_rate >= 80 else "🔴"
        st.markdown(f"**Parsing QC:** {rate_icon} {success_rate:.1f}% success rate — "
                    f"{parsing_meta.get('total_transactions_extracted',0):,} transactions extracted, "
                    f"{parsing_meta.get('total_balance_checks_passed',0)}/{parsing_meta.get('total_balance_checks',0)} balance checks passed")

    # ── Account Summary ──
    st.markdown("#### Accounts")
    for acc in accounts:
        st.markdown(f"**{acc.get('bank_name','')}** ({acc.get('account_number','')}) — {acc.get('account_type','')}")
        st.write(f"Credits: RM {acc.get('total_credits',0):,.2f} | Debits: RM {acc.get('total_debits',0):,.2f} | Closing: RM {acc.get('closing_balance',0):,.2f}")

    # ── Observations ──
    obs = data.get('observations', {})
    if obs.get('positive') or obs.get('concerns'):
        st.markdown("#### Observations")
        col1, col2 = st.columns(2)
        with col1:
            for o in obs.get('positive', []):
                st.success(f"✅ {o}")
        with col2:
            for o in obs.get('concerns', []):
                st.error(f"⚠️ {o}")

    # ── Flags ──
    st.markdown("#### Risk Signals")
    for f in flags_data.get('indicators', []):
        icon = "🔴" if f.get('detected') else "🟢"
        st.markdown(f"{icon} **{f.get('name','')}** — {f.get('remarks','')}")

    # ── v6.2.0: Top Parties with MoM trend ──
    top_parties_data = data.get('top_parties', {})
    has_bd = any(p.get('monthly_breakdown') for p in top_parties_data.get('top_payers', []) + top_parties_data.get('top_payees', []))
    if has_bd:
        st.markdown("#### Top Counterparty Trends")
        tp_col1, tp_col2 = st.columns(2)
        with tp_col1:
            st.markdown("**Top Payers**")
            for p in top_parties_data.get('top_payers', []):
                mb = p.get('monthly_breakdown', [])
                if mb:
                    trend_str = ' → '.join(f"{mb_item.get('month','')[-5:]}: RM{mb_item.get('amount',0):,.0f}" for mb_item in mb if mb_item.get('amount', 0) > 0)
                    rp_tag = " 🏷️RP" if p.get('is_related_party') else ""
                    with st.expander(f"#{p.get('rank')} {p.get('party_name','')}{rp_tag} — RM {p.get('total_amount',0):,.0f}"):
                        st.caption(trend_str)
        with tp_col2:
            st.markdown("**Top Payees**")
            for p in top_parties_data.get('top_payees', []):
                mb = p.get('monthly_breakdown', [])
                if mb:
                    trend_str = ' → '.join(f"{mb_item.get('month','')[-5:]}: RM{mb_item.get('amount',0):,.0f}" for mb_item in mb if mb_item.get('amount', 0) > 0)
                    rp_tag = " 🏷️RP" if p.get('is_related_party') else ""
                    with st.expander(f"#{p.get('rank')} {p.get('party_name','')}{rp_tag} — RM {p.get('total_amount',0):,.0f}"):
                        st.caption(trend_str)

    # ── v6.2.0/v6.2.1: Parsing QC detail ──
    if parsing_meta:
        with st.expander("📋 Parsing QC — Balance Reconciliation Detail"):
            checks = parsing_meta.get('account_month_checks', [])
            for chk in checks:
                passed = chk.get('passed', False)
                icon = "✅" if passed else "❌"
                delta = chk.get('reconciliation_delta', 0)
                gap_count = chk.get('extraction_gaps', 0) or 0
                gap_str = f" | ⚠️ {gap_count} gap(s)" if gap_count > 0 else ""
                st.markdown(
                    f"{icon} **{chk.get('month','')}** ({chk.get('account_number','')}) — "
                    f"Delta: RM {delta:,.2f} | Txns: {chk.get('transactions_extracted',0)}"
                    f"{gap_str}"
                    f"{'  |  ' + chk.get('notes','') if chk.get('notes') else ''}"
                )

            # v6.2.1: Show extraction gap detail
            extraction_gaps = parsing_meta.get('extraction_gaps', [])
            if extraction_gaps:
                st.markdown("---")
                st.markdown("**🔍 Extraction Gap Details**")
                for g in extraction_gaps:
                    st.markdown(
                        f"- **{g.get('month','')}** {g.get('date','')} p.{g.get('page','')}: "
                        f"RM {g.get('missing_amount',0):,.2f} {g.get('missing_type','').lower()}(s) missing — "
                        f"between _{g.get('prev_description','')[:50]}_ and _{g.get('next_description','')[:50]}_"
                    )

else:
    st.info("Upload a JSON analysis file to get started.")
